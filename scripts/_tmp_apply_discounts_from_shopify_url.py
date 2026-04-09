"""Apply discount tags from Excel using Shopify URL product IDs.

Rules:
- Extract Shopify product ID from `Shopify URL` like .../admin/products/<id>
- Parse discount value from `Discount ` column
- Find normalized doc by Shopify ID (channels.shopify.shopify_id OR shopify_id)
- Remove any existing discount_* tags, then add new discount_<value> tag
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymongo import MongoClient

from app.config import settings

EXCEL_PATH = Path("/Users/administrator/Downloads/Gallery1880_Shipping & Discount Profile.xlsx")
SHOPIFY_URL_COL = "Shopify URL"
DISCOUNT_COL = "Discount"

URL_ID_RE = re.compile(r"/admin/products/(\d+)")
DISCOUNT_TAG_RE = re.compile(r"^discount_\d+$", re.IGNORECASE)
DISCOUNT_VALUE_RE = re.compile(r"(\d+)")


def to_discount_int(raw: Any) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    text = str(raw).strip()
    if not text:
        return None
    match = DISCOUNT_VALUE_RE.search(text)
    if not match:
        return None
    return int(match.group(1))


def extract_shopify_product_id(url: Any) -> int | None:
    if not url:
        return None
    text = str(url).strip()
    match = URL_ID_RE.search(text)
    if not match:
        return None
    return int(match.group(1))


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_to_idx = {h: i for i, h in enumerate(headers)}

    if SHOPIFY_URL_COL not in header_to_idx:
        raise ValueError(f"Missing required column: {SHOPIFY_URL_COL}")
    if DISCOUNT_COL not in header_to_idx:
        # Fallback for slight naming drift in spreadsheet exports.
        alt = next((h for h in headers if h.lower().startswith("discount")), None)
        if alt:
            header_to_idx[DISCOUNT_COL] = header_to_idx[alt]
        else:
            raise ValueError(f"Missing required column: {DISCOUNT_COL}")

    idx_url = header_to_idx[SHOPIFY_URL_COL]
    idx_discount = header_to_idx[DISCOUNT_COL]

    rows_total = 0
    rows_with_url = 0
    rows_with_discount = 0
    rows_valid = 0

    # product_id -> discount_value; if duplicates conflict, keep the last seen and report.
    mapping: dict[int, int] = {}
    conflicts: list[tuple[int, int, int]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_total += 1
        raw_url = row[idx_url] if idx_url < len(row) else None
        raw_discount = row[idx_discount] if idx_discount < len(row) else None

        pid = extract_shopify_product_id(raw_url)
        if pid is None:
            continue
        rows_with_url += 1

        dval = to_discount_int(raw_discount)
        if dval is None:
            continue
        rows_with_discount += 1

        if pid in mapping and mapping[pid] != dval:
            conflicts.append((pid, mapping[pid], dval))
        mapping[pid] = dval
        rows_valid += 1

    client = MongoClient(settings.MONGO_URI)
    db = client[settings.MONGO_DB]
    col = db.product_normalized

    matched_docs = 0
    updated_docs = 0
    unchanged_docs = 0
    missing_docs = 0

    for pid, dval in mapping.items():
        query = {
            "$or": [
                {"channels.shopify.shopify_id": pid},
                {"channels.shopify.shopify_id": str(pid)},
                {"shopify_id": pid},
                {"shopify_id": str(pid)},
            ]
        }
        doc = col.find_one(query, {"_id": 1, "tags": 1})
        if not doc:
            missing_docs += 1
            continue

        matched_docs += 1
        tags = doc.get("tags") or []
        if not isinstance(tags, list):
            tags = []

        filtered = [t for t in tags if not (isinstance(t, str) and DISCOUNT_TAG_RE.match(t.strip()))]
        new_tag = f"discount_{dval}"

        if new_tag not in filtered:
            filtered.append(new_tag)

        if filtered == tags:
            unchanged_docs += 1
            continue

        col.update_one({"_id": doc["_id"]}, {"$set": {"tags": filtered}})
        updated_docs += 1

    print(f"excel_path={EXCEL_PATH}")
    print(f"mongo_db={settings.MONGO_DB}")
    print(f"rows_total={rows_total}")
    print(f"rows_with_shopify_url={rows_with_url}")
    print(f"rows_with_discount={rows_with_discount}")
    print(f"rows_valid_for_mapping={rows_valid}")
    print(f"unique_shopify_products_from_excel={len(mapping)}")
    print(f"conflicting_duplicate_rows={len(conflicts)}")
    print(f"matched_docs={matched_docs}")
    print(f"updated_docs={updated_docs}")
    print(f"unchanged_docs={unchanged_docs}")
    print(f"missing_docs={missing_docs}")

    if conflicts:
        print("conflict_examples=")
        for pid, oldv, newv in conflicts[:10]:
            print(f"  shopify_id={pid} old={oldv} new={newv}")


if __name__ == "__main__":
    main()
