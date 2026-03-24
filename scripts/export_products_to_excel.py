"""Export products from MongoDB to an Excel (.xlsx) file.

Columns:
- Product ID
- 1 Image URL
- Shopify URL
- Price
- Availability (Available / Not Available)
- Concatenated Dimensions
- Weights

By default, reads from the `product_normalized` collection.

Usage examples:
  python -m scripts.export_products_to_excel --out products.xlsx
  python -m scripts.export_products_to_excel --limit 100
  python -m scripts.export_products_to_excel --query '{"shopify_id": {"$exists": true}}'

Environment variables (loaded from .env if present):
- MONGO_URI (required)
- MONGO_DB (required)
- SHOPIFY_STORE_URL (optional; used to build admin product URLs)
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime
from typing import Any

from dotenv import load_dotenv
from pymongo import MongoClient

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing dependency `openpyxl`. Install it with: pip install openpyxl"
    ) from e


HEADERS = [
    "Product ID",
    "1 Image URL",
    "Shopify URL",
    "Price",
    "Availability",
    "Concatinated Dimensions",
    "Weights",
]


def _num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace("$", "")
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _first_image_url(doc: dict[str, Any]) -> str | None:
    images = doc.get("images") or []
    if isinstance(images, (list, tuple)) and images:
        first = images[0]
        return str(first) if first else None
    return None


def _shopify_admin_url(store_url: str | None, shopify_id: Any) -> str | None:
    if not store_url or not shopify_id:
        return None
    store_url = str(store_url).strip()
    if not store_url:
        return None
    # store_url should be like "my-store.myshopify.com" (no scheme)
    if store_url.startswith("http://"):
        store_url = store_url[len("http://") :]
    if store_url.startswith("https://"):
        store_url = store_url[len("https://") :]
    store_url = store_url.rstrip("/")
    return f"https://{store_url}/admin/products/{shopify_id}"


def _availability(quantity: Any) -> str:
    q = _num(quantity)
    return "Available" if (q is not None and q > 0) else "Not Available"


def _concat_dimensions(doc: dict[str, Any]) -> str | None:
    package = doc.get("package") or {}
    if not isinstance(package, dict):
        return None
    dims = package.get("dimensions") or {}
    if not isinstance(dims, dict):
        return None

    parts: list[str] = []
    units: list[str] = []
    for key in ("length", "width", "height"):
        m = dims.get(key)
        if not isinstance(m, dict):
            continue
        v = _num(m.get("value"))
        if v is None:
            continue
        unit = (m.get("unit") or "").strip()
        parts.append(f"{v:g}")
        if unit:
            units.append(unit)

    if not parts:
        return None

    # If all units match, format like "10 x 5 x 3 in"; otherwise omit the unit.
    unit_out = None
    if units and len(set(units)) == 1 and len(units) == len(parts):
        unit_out = units[0]

    s = " x ".join(parts)
    if unit_out:
        s = f"{s} {unit_out}"
    return s


def _weight_string(doc: dict[str, Any]) -> str | None:
    package = doc.get("package") or {}
    if not isinstance(package, dict):
        return None
    weight = package.get("weight") or {}
    if not isinstance(weight, dict):
        return None

    def _piece(m: Any) -> str | None:
        if not isinstance(m, dict):
            return None
        v = _num(m.get("value"))
        if v is None:
            return None
        unit = (m.get("unit") or "").strip()
        if unit:
            return f"{v:g} {unit}"
        return f"{v:g}"

    major = _piece(weight.get("major"))
    minor = _piece(weight.get("minor"))

    if major and minor:
        return f"{major} {minor}"
    return major or minor


def _parse_query(query_str: str | None) -> dict[str, Any]:
    if not query_str:
        return {}
    try:
        q = json.loads(query_str)
    except json.JSONDecodeError as e:
        raise SystemExit(f"Invalid --query JSON: {e}")
    if not isinstance(q, dict):
        raise SystemExit("--query must be a JSON object")
    return q


def export_to_excel(
    *,
    mongo_uri: str,
    mongo_db: str,
    collection: str,
    out_path: str,
    query: dict[str, Any],
    limit: int | None,
    shopify_store_url: str | None,
) -> int:
    client = MongoClient(mongo_uri)
    col = client[mongo_db][collection]

    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    projection = {
        "_id": 1,
        "sku": 1,
        "images": 1,
        "shopify_id": 1,
        "price": 1,
        "quantity": 1,
        "package": 1,
    }

    cursor = col.find(query or {}, projection)
    if limit is not None:
        cursor = cursor.limit(int(limit))

    count = 0
    for doc in cursor:
        product_id = doc.get("sku") or doc.get("_id")
        image_url = _first_image_url(doc)
        shopify_url = _shopify_admin_url(shopify_store_url, doc.get("shopify_id"))
        price = _num(doc.get("price"))
        availability = _availability(doc.get("quantity"))
        dims = _concat_dimensions(doc)
        weight = _weight_string(doc)

        ws.append(
            [
                str(product_id) if product_id is not None else None,
                image_url,
                shopify_url,
                price,
                availability,
                dims,
                weight,
            ]
        )
        count += 1

    # Basic column widths for readability
    widths = [
        18,  # Product ID
        60,  # Image URL
        45,  # Shopify URL
        12,  # Price
        14,  # Availability
        24,  # Dimensions
        18,  # Weight
    ]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    wb.save(out_path)
    return count


def main() -> None:
    load_dotenv()

    parser = argparse.ArgumentParser(description="Export MongoDB products to Excel")
    parser.add_argument(
        "--out",
        default=None,
        help="Output .xlsx file path (default: export_products_YYYYMMDD_HHMMSS.xlsx)",
    )
    parser.add_argument(
        "--collection",
        default=os.getenv("MONGO_COLLECTION", "product_normalized"),
        help="Mongo collection to read (default: product_normalized)",
    )
    parser.add_argument(
        "--query",
        default=None,
        help="Mongo find() query as JSON string (default: none)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Max number of documents to export (default: no limit)",
    )
    parser.add_argument(
        "--mongo-uri",
        default=os.getenv("MONGO_URI"),
        help="Mongo URI (default: env MONGO_URI)",
    )
    parser.add_argument(
        "--mongo-db",
        default=os.getenv("MONGO_DB"),
        help="Mongo DB name (default: env MONGO_DB)",
    )
    parser.add_argument(
        "--shopify-store-url",
        default=(
            os.getenv("SHOPIFY_STORE_URL")
            or os.getenv("SHOPIFY_STORE_URL_PROD")
            or os.getenv("SHOPIFY_SHOP")
        ),
        help="Shopify store hostname used to build admin URLs (default: env SHOPIFY_STORE_URL)",
    )

    args = parser.parse_args()

    mongo_uri = args.mongo_uri
    mongo_db = args.mongo_db

    if not mongo_uri:
        raise SystemExit("Missing Mongo connection info: set MONGO_URI or pass --mongo-uri")
    if not mongo_db:
        raise SystemExit("Missing Mongo DB name: set MONGO_DB or pass --mongo-db")

    out_path = args.out
    if not out_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"export_products_{ts}.xlsx"

    query = _parse_query(args.query)

    count = export_to_excel(
        mongo_uri=mongo_uri,
        mongo_db=mongo_db,
        collection=args.collection,
        out_path=out_path,
        query=query,
        limit=args.limit,
        shopify_store_url=args.shopify_store_url,
    )

    print(f"Wrote {count} rows to {out_path}")


if __name__ == "__main__":
    main()
