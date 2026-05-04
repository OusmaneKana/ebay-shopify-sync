from __future__ import annotations

import argparse
import asyncio
import copy
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from app.config import settings
from app.database.mongo import close_mongo_client, db
from app.services.etsy_auth_service import get_valid_token as get_valid_etsy_token

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REPORT_PATH = ROOT / "logs" / "zero_qty_exact_title_matches_20260426_024121.xlsx"
DEFAULT_SHEET_NAME = "zero_qty_title_matches"
TOP_LEVEL_COPY_FIELDS = (
    "compare_at_price",
    "discount_percent",
    "sale_active",
    "sale_start",
    "sale_end",
)
SHOPIFY_CHANNEL_COPY_FIELDS = (
    "compare_at_price",
    "discount_percent",
    "sale_active",
    "sale_start",
    "sale_end",
)


class MigrationError(RuntimeError):
    pass


@dataclass(slots=True)
class ReportRow:
    row_index: int
    old_sku: str
    old_qty: int | None
    title: str
    new_sku: str
    new_qty: int | None
    old_etsy_listing_id: str | None
    old_etsy_state: str | None
    new_etsy_listing_id: str | None
    new_etsy_state: str | None
    old_updated_at: Any
    new_updated_at: Any


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if value == "":
        return True
    if isinstance(value, dict) and not value:
        return True
    return False


def _money_to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        amount = value.get("amount")
        divisor = value.get("divisor")
        if amount is None or divisor in (None, 0):
            raise MigrationError(f"invalid_money_value={value!r}")
        return float(amount) / float(divisor)
    raise MigrationError(f"unsupported_price_value={value!r}")


def _load_report_rows(report_path: Path, sheet_name: str) -> list[ReportRow]:
    workbook = load_workbook(report_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise MigrationError(f"missing_sheet={sheet_name}")

    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    header_map = {_normalize_header(value): index for index, value in enumerate(headers)}
    required_headers = {
        "old_sku",
        "old_qty",
        "title",
        "new_sku",
        "new_qty",
        "old_etsy_listing_id",
        "old_etsy_state",
        "new_etsy_listing_id",
        "new_etsy_state",
        "old_updated_at",
        "new_updated_at",
    }
    missing_headers = sorted(required_headers - set(header_map))
    if missing_headers:
        raise MigrationError(f"missing_headers={','.join(missing_headers)}")

    parsed_rows: list[ReportRow] = []
    for row_index, values in enumerate(rows, start=2):
        old_sku = str(values[header_map["old_sku"]] or "").strip()
        new_sku = str(values[header_map["new_sku"]] or "").strip()
        title = str(values[header_map["title"]] or "").strip()
        if not old_sku or not new_sku or not title:
            continue

        parsed_rows.append(
            ReportRow(
                row_index=row_index,
                old_sku=old_sku,
                old_qty=values[header_map["old_qty"]],
                title=title,
                new_sku=new_sku,
                new_qty=values[header_map["new_qty"]],
                old_etsy_listing_id=_optional_str(values[header_map["old_etsy_listing_id"]]),
                old_etsy_state=_optional_str(values[header_map["old_etsy_state"]]),
                new_etsy_listing_id=_optional_str(values[header_map["new_etsy_listing_id"]]),
                new_etsy_state=_optional_str(values[header_map["new_etsy_state"]]),
                old_updated_at=values[header_map["old_updated_at"]],
                new_updated_at=values[header_map["new_updated_at"]],
            )
        )

    return parsed_rows


def _optional_str(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _build_one_to_one_indexes(rows: list[ReportRow]) -> tuple[dict[str, int], dict[str, int]]:
    old_counts: dict[str, int] = {}
    new_counts: dict[str, int] = {}
    for row in rows:
        old_counts[row.old_sku] = old_counts.get(row.old_sku, 0) + 1
        new_counts[row.new_sku] = new_counts.get(row.new_sku, 0) + 1
    return old_counts, new_counts


def _is_safe_report_row(row: ReportRow, old_counts: dict[str, int], new_counts: dict[str, int]) -> tuple[bool, str]:
    if old_counts.get(row.old_sku) != 1:
        return False, "old_sku_not_one_to_one"
    if new_counts.get(row.new_sku) != 1:
        return False, "new_sku_not_one_to_one"
    if not row.old_etsy_listing_id:
        return False, "old_listing_missing"
    if row.new_etsy_listing_id:
        return False, "new_listing_already_linked"
    if row.old_sku == row.new_sku:
        return False, "same_sku"
    return True, "ok"


async def _fetch_doc_by_sku(sku: str) -> dict[str, Any] | None:
    return await db.product_normalized.find_one({"sku": sku})


def _channel_etsy(doc: dict[str, Any]) -> dict[str, Any] | None:
    return ((doc.get("channels") or {}).get("etsy")) if doc else None


def _channel_shopify(doc: dict[str, Any]) -> dict[str, Any] | None:
    return ((doc.get("channels") or {}).get("shopify")) if doc else None


def _build_new_doc_set_ops(old_doc: dict[str, Any], new_doc: dict[str, Any]) -> dict[str, Any]:
    set_ops: dict[str, Any] = {}

    for field in TOP_LEVEL_COPY_FIELDS:
        old_value = old_doc.get(field)
        new_value = new_doc.get(field)
        if not _is_missing(old_value) and _is_missing(new_value):
            set_ops[field] = old_value

    old_shopify = _channel_shopify(old_doc) or {}
    new_shopify = _channel_shopify(new_doc) or {}
    if new_shopify:
        for field in SHOPIFY_CHANNEL_COPY_FIELDS:
            old_value = old_shopify.get(field)
            new_value = new_shopify.get(field)
            if not _is_missing(old_value) and _is_missing(new_value):
                set_ops[f"channels.shopify.{field}"] = old_value

    old_etsy = _channel_etsy(old_doc)
    new_etsy = _channel_etsy(new_doc)
    if old_etsy and _is_missing(new_etsy):
        set_ops["channels.etsy"] = copy.deepcopy(old_etsy)

    return set_ops


def _validate_docs(row: ReportRow, old_doc: dict[str, Any] | None, new_doc: dict[str, Any] | None) -> tuple[bool, str]:
    if not old_doc:
        return False, "old_doc_missing"
    if not new_doc:
        return False, "new_doc_missing"
    if old_doc.get("_id") == new_doc.get("_id"):
        return False, "same_document"
    if str(old_doc.get("title") or "").strip() != row.title:
        return False, "old_title_mismatch"
    if str(new_doc.get("title") or "").strip() != row.title:
        return False, "new_title_mismatch"
    if int(old_doc.get("quantity") or 0) != 0:
        return False, "old_quantity_not_zero"
    if int(new_doc.get("quantity") or 0) <= 0:
        return False, "new_quantity_not_positive"

    old_etsy = _channel_etsy(old_doc) or {}
    new_etsy = _channel_etsy(new_doc) or {}
    old_listing_id = _optional_str(old_etsy.get("listing_id"))
    new_listing_id = _optional_str(new_etsy.get("listing_id"))
    if old_listing_id != row.old_etsy_listing_id:
        return False, "old_listing_id_mismatch"
    if new_listing_id:
        return False, "new_doc_already_has_etsy"
    shop_id = _optional_str(old_etsy.get("shop_id"))
    if not shop_id:
        return False, "old_shop_id_missing"
    return True, "ok"


async def _resolve_etsy_auth_headers() -> dict[str, str]:
    try:
        token = await get_valid_etsy_token()
    except ValueError as exc:
        raise MigrationError("missing_etsy_token") from exc

    if settings.ETSY_CLIENT_ID and settings.ETSY_CLIENT_SECRET:
        api_key = f"{settings.ETSY_CLIENT_ID}:{settings.ETSY_CLIENT_SECRET}"
    else:
        api_key = settings.ETSY_CLIENT_ID
    if not api_key:
        raise MigrationError("missing_etsy_api_key")

    return {
        "Authorization": f"Bearer {token}",
        "x-api-key": api_key,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def _build_inventory_update_payload(inventory: dict[str, Any], new_sku: str, target_quantity: int) -> dict[str, Any]:
    products = inventory.get("products") or []
    if not products:
        raise MigrationError("etsy_inventory_missing_products")

    payload_products: list[dict[str, Any]] = []
    for product in products:
        offerings = product.get("offerings") or []
        if not offerings:
            raise MigrationError("etsy_inventory_product_missing_offerings")

        payload_offerings: list[dict[str, Any]] = []
        positive_quantity_found = False
        for offering in offerings:
            quantity = int(offering.get("quantity") or 0)
            if quantity > 0:
                positive_quantity_found = True
            payload_offerings.append(
                {
                    "price": _money_to_float(offering.get("price")),
                    "quantity": quantity,
                    "is_enabled": bool(offering.get("is_enabled")),
                    "readiness_state_id": offering.get("readiness_state_id"),
                }
            )

        if not positive_quantity_found:
            if target_quantity <= 0:
                raise MigrationError("etsy_inventory_requires_positive_quantity")
            first_enabled_index = next(
                (index for index, offering in enumerate(payload_offerings) if offering["is_enabled"]),
                0,
            )
            payload_offerings[first_enabled_index]["quantity"] = target_quantity

        payload_product: dict[str, Any] = {
            "sku": new_sku,
            "offerings": payload_offerings,
        }
        property_values = product.get("property_values") or []
        if property_values:
            payload_product["property_values"] = copy.deepcopy(property_values)

        payload_products.append(payload_product)

    return {
        "products": payload_products,
        "price_on_property": copy.deepcopy(inventory.get("price_on_property") or []),
        "quantity_on_property": copy.deepcopy(inventory.get("quantity_on_property") or []),
        "sku_on_property": copy.deepcopy(inventory.get("sku_on_property") or []),
        "readiness_state_on_property": copy.deepcopy(inventory.get("readiness_state_on_property") or []),
    }


async def _update_etsy_listing_sku(
    *, listing_id: str, new_sku: str, target_quantity: int, headers: dict[str, str]
) -> dict[str, Any]:
    async with httpx.AsyncClient(timeout=45.0) as client:
        get_response = await client.get(
            f"https://openapi.etsy.com/v3/application/listings/{listing_id}/inventory",
            headers=headers,
            params={"show_deleted": "true"},
        )
        if get_response.status_code >= 400:
            raise MigrationError(
                f"etsy_inventory_get_failed status={get_response.status_code} body={get_response.text}"
            )

        inventory_payload = get_response.json() if get_response.content else {}
        update_payload = _build_inventory_update_payload(inventory_payload, new_sku, target_quantity)

        put_response = await client.put(
            f"https://openapi.etsy.com/v3/application/listings/{listing_id}/inventory",
            headers=headers,
            content=json.dumps(update_payload),
        )
        if put_response.status_code >= 400:
            raise MigrationError(
                f"etsy_inventory_put_failed status={put_response.status_code} body={put_response.text}"
            )

        return put_response.json() if put_response.content else {}


async def _apply_row(
    row: ReportRow,
    *,
    etsy_headers: dict[str, str] | None,
    apply: bool,
    report_path: Path,
    skip_etsy_update: bool,
) -> dict[str, Any]:
    old_doc = await _fetch_doc_by_sku(row.old_sku)
    new_doc = await _fetch_doc_by_sku(row.new_sku)
    valid, reason = _validate_docs(row, old_doc, new_doc)
    if not valid:
        return {"row": row.row_index, "status": "skipped", "reason": reason}

    assert old_doc is not None
    assert new_doc is not None
    old_etsy = _channel_etsy(old_doc) or {}
    listing_id = _optional_str(old_etsy.get("listing_id"))
    shop_id = _optional_str(old_etsy.get("shop_id"))
    if not listing_id or not shop_id:
        return {"row": row.row_index, "status": "skipped", "reason": "old_etsy_metadata_incomplete"}
    stored_listing_id = old_etsy.get("listing_id")

    set_ops = _build_new_doc_set_ops(old_doc, new_doc)
    if "channels.etsy" not in set_ops:
        return {"row": row.row_index, "status": "skipped", "reason": "new_doc_not_missing_etsy"}

    new_etsy_channel = copy.deepcopy(set_ops["channels.etsy"])
    new_etsy_channel["sku"] = row.new_sku
    set_ops["channels.etsy"] = new_etsy_channel

    migration_now = _utc_now()
    set_ops["migrations.etsy_sku_relink"] = {
        "row_index": row.row_index,
        "from_sku": row.old_sku,
        "to_sku": row.new_sku,
        "listing_id": listing_id,
        "report_path": str(report_path),
        "applied_at": migration_now,
    }
    old_set_ops = {
        "migrations.etsy_sku_relink": {
            "row_index": row.row_index,
            "from_sku": row.old_sku,
            "to_sku": row.new_sku,
            "listing_id": listing_id,
            "moved_at": migration_now,
        }
    }
    old_unset_ops = {"channels.etsy": ""}

    copied_fields = sorted(key for key in set_ops if key not in {"channels.etsy", "migrations.etsy_sku_relink"})
    result: dict[str, Any] = {
        "row": row.row_index,
        "status": "planned" if not apply else "applied",
        "old_sku": row.old_sku,
        "new_sku": row.new_sku,
        "listing_id": listing_id,
        "copied_fields": copied_fields,
        "moved_etsy": True,
        "etsy_updated": False,
    }

    if not apply:
        return result

    if not skip_etsy_update:
        if not etsy_headers:
            raise MigrationError("etsy_headers_missing_for_apply")
        await _update_etsy_listing_sku(
            listing_id=listing_id,
            new_sku=row.new_sku,
            target_quantity=int(new_doc.get("quantity") or 0),
            headers=etsy_headers,
        )
        result["etsy_updated"] = True

    await db.product_normalized.update_one(
        {"_id": new_doc["_id"], "channels.etsy": {"$exists": False}},
        {"$set": set_ops},
    )
    await db.product_normalized.update_one(
        {"_id": old_doc["_id"], "channels.etsy.listing_id": stored_listing_id},
        {"$set": old_set_ops, "$unset": old_unset_ops},
    )
    return result


async def _run(args: argparse.Namespace) -> int:
    report_path = Path(args.report).expanduser().resolve()
    rows = _load_report_rows(report_path, args.sheet)
    old_counts, new_counts = _build_one_to_one_indexes(rows)

    filtered_rows: list[ReportRow] = []
    skipped_rows: list[dict[str, Any]] = []
    for row in rows:
        safe, reason = _is_safe_report_row(row, old_counts, new_counts)
        if not safe:
            skipped_rows.append({"row": row.row_index, "status": "skipped", "reason": reason})
            continue
        if args.old_sku and row.old_sku != args.old_sku:
            continue
        if args.new_sku and row.new_sku != args.new_sku:
            continue
        filtered_rows.append(row)

    if args.limit:
        filtered_rows = filtered_rows[: args.limit]

    etsy_headers = None
    if args.apply and not args.skip_etsy_update:
        etsy_headers = await _resolve_etsy_auth_headers()

    results: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row in filtered_rows:
        try:
            result = await _apply_row(
                row,
                etsy_headers=etsy_headers,
                apply=args.apply,
                report_path=report_path,
                skip_etsy_update=args.skip_etsy_update,
            )
            results.append(result)
        except Exception as exc:
            errors.append({"row": row.row_index, "status": "error", "reason": str(exc)})

    summary = {
        "report": str(report_path),
        "sheet": args.sheet,
        "apply": args.apply,
        "skip_etsy_update": args.skip_etsy_update,
        "safe_candidates_considered": len(filtered_rows),
        "precheck_skipped": len(skipped_rows),
        "planned_or_applied": len(results),
        "errors": len(errors),
        "results": results,
        "skipped": skipped_rows,
        "error_rows": errors,
    }
    print(json.dumps(summary, default=str, indent=2))
    return 1 if errors else 0


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Relink Etsy listings from old zero-quantity SKUs to new active SKUs using the exact-title report."
    )
    parser.add_argument("--report", default=str(DEFAULT_REPORT_PATH), help="Path to the workbook report.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet name inside the workbook.")
    parser.add_argument("--apply", action="store_true", help="Apply Mongo and Etsy changes. Default is dry-run.")
    parser.add_argument(
        "--skip-etsy-update",
        action="store_true",
        help="Skip the remote Etsy inventory SKU update. Mongo relink still requires --apply.",
    )
    parser.add_argument("--old-sku", help="Only process a specific old SKU.")
    parser.add_argument("--new-sku", help="Only process a specific new SKU.")
    parser.add_argument("--limit", type=int, help="Limit the number of safe candidate rows processed.")
    return parser


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()
    try:
        raise SystemExit(asyncio.run(_run(args)))
    finally:
        close_mongo_client()


if __name__ == "__main__":
    main()