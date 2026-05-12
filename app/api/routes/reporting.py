import re
import json
import copy
import asyncio
import mimetypes
import math
from pathlib import Path
from datetime import datetime, timezone
from typing import Any
from urllib.parse import urlparse

import httpx
from openai import OpenAI
from openpyxl import load_workbook
from fastapi import APIRouter, Query, HTTPException, Body

from app.database.mongo import db
from app.config import settings
from app.services.etsy_auth_service import get_valid_token as get_valid_etsy_token
from app.shopify.client import ShopifyClient
router = APIRouter()

MATCH_REPORT_PATH = Path("logs/etsy_title_match_analysis.json")
REVIEW_COLLECTION = "etsy_match_review"
EXCEL_REVIEW_PATTERNS = [
    "active_etsy_no_ebay_match_with_mongo_candidates_and_shopify_link_*.xlsx",
    "active_etsy_no_ebay_match_with_mongo_candidates_*.xlsx",
    "active_etsy_no_ebay_match_*.xlsx",
]
ETSY_BASE_URL = "https://openapi.etsy.com/v3/application"
ETSY_REQUIRED_CREATE_FIELDS = [
    ("title", "Title"),
    ("description", "Description"),
    ("price", "Price"),
    ("quantity", "Quantity"),
    ("type", "Listing Type"),
    ("who_made", "Who Made"),
    ("when_made", "When Made"),
    ("taxonomy_id", "Taxonomy ID"),
    ("shipping_profile_id", "Shipping Profile ID (physical only)"),
    ("return_policy_id", "Return Policy ID"),
]

ETSY_TAXONOMY_OPENAI_MODEL = getattr(settings, "OPENAI_MODEL_ETSY_TAXONOMY", "gpt-4.1-mini")
ETSY_TAG_OPENAI_MODEL = getattr(settings, "OPENAI_MODEL_ETSY_TAGS", ETSY_TAXONOMY_OPENAI_MODEL)
ETSY_SEO_OPENAI_MODEL = getattr(settings, "OPENAI_MODEL_ETSY_SEO", ETSY_TAXONOMY_OPENAI_MODEL)
ETSY_TAXONOMY_PREFILTER_LIMIT = 40
ETSY_TAXONOMY_RESULT_LIMIT = 8
ETSY_TAXONOMY_CACHE_TTL_SECONDS = 60 * 60 * 12
_ETSY_BUYER_TAXONOMY_CACHE: dict[str, Any] = {"fetched_at": None, "nodes": None}
_ETSY_READINESS_CACHE: dict[str, Any] = {"by_shop": {}}
ETSY_TAG_MAX_COUNT = 13
ETSY_TAG_MAX_LENGTH = 20
ETSY_TITLE_MAX_LENGTH = 140

ETSY_ALLOWED_LISTING_TYPES = {"physical", "digital"}

ETSY_ALLOWED_WHO_MADE = {"i_did", "collective", "someone_else"}
ETSY_ALLOWED_WHEN_MADE = {
    "made_to_order",
    "2020_2026",
    "2010_2019",
    "2007_2009",
    "2000_2006",
    "before_2007",
    "1990s",
    "1980s",
    "1970s",
    "1960s",
    "1950s",
    "1940s",
    "1930s",
    "1920s",
    "1910s",
    "1900s",
    "1800s",
    "1700s",
    "before_1700",
}
ETSY_ALLOWED_TAG_CHARS_RE = re.compile(r"[^A-Za-z0-9 '\-™©®]+")

# Bulk publishing configuration
ETSY_BULK_MIN_TAXONOMY_CONFIDENCE = 0.5
ETSY_BULK_MAX_ITEMS_PER_RUN = 500
ETSY_BULK_VALIDATE_CONCURRENCY = 8
ETSY_BULK_VALIDATE_BATCH_DELAY_SECONDS = 0.0
BULK_REPORT_COLLECTION = "etsy_bulk_reports"
_ETSY_BULK_REPORT: dict[str, Any] = {
    "session_id": None,
    "validated_at": None,
    "created_at": None,
    "validation_result": None,
    "creation_result": None,
}

# Rate limiting configuration
# OpenAI: ~3 requests/min for gpt-4, ~60 requests/min for gpt-3.5
# Etsy: 10 requests/second per API key (1000 per 100 seconds)
OPENAI_MAX_CONCURRENT_REQUESTS = 3  # Faster throughput while still bounded
OPENAI_REQUEST_DELAY_SECONDS = 0.1  # Minimal pacing between OpenAI calls
ETSY_MAX_CONCURRENT_REQUESTS = 5    # 5 concurrent requests (well below 10/sec limit)
ETSY_REQUEST_DELAY_SECONDS = 0.1    # 100ms delay between requests

# Semaphores for rate limiting (initialized once)
_openai_semaphore: asyncio.Semaphore | None = None
_etsy_semaphore: asyncio.Semaphore | None = None


def _get_openai_semaphore() -> asyncio.Semaphore:
    """Get or create the OpenAI rate limit semaphore."""
    global _openai_semaphore
    if _openai_semaphore is None:
        _openai_semaphore = asyncio.Semaphore(OPENAI_MAX_CONCURRENT_REQUESTS)
    return _openai_semaphore


def _get_etsy_semaphore() -> asyncio.Semaphore:
    """Get or create the Etsy rate limit semaphore."""
    global _etsy_semaphore
    if _etsy_semaphore is None:
        _etsy_semaphore = asyncio.Semaphore(ETSY_MAX_CONCURRENT_REQUESTS)
    return _etsy_semaphore


def _ensure_utc(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _rx(text: str) -> dict:
    # Escape to avoid accidental regex patterns coming from user input
    return {"$regex": re.escape(text), "$options": "i"}


def _load_match_report() -> dict:
    if not MATCH_REPORT_PATH.exists():
        raise HTTPException(status_code=404, detail="Match report not found. Export analysis first.")
    with MATCH_REPORT_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _bulk_reports_collection():
    return db[BULK_REPORT_COLLECTION]


async def _persist_bulk_validation_report(
    *,
    session_id: str,
    validated_at: datetime,
    validation_result: dict[str, Any],
) -> None:
    await _bulk_reports_collection().update_one(
        {"session_id": session_id},
        {
            "$set": {
                "session_id": session_id,
                "validated_at": validated_at,
                "validation_result": validation_result,
                "creation_result": None,
                "created_at": None,
                "updated_at": _now_utc(),
            },
            "$setOnInsert": {
                "created_doc_at": validated_at,
            },
        },
        upsert=True,
    )


async def _persist_bulk_creation_report(
    *,
    session_id: str,
    created_at: datetime,
    creation_result: dict[str, Any],
) -> None:
    await _bulk_reports_collection().update_one(
        {"session_id": session_id},
        {
            "$set": {
                "created_at": created_at,
                "creation_result": creation_result,
                "updated_at": _now_utc(),
            }
        },
        upsert=True,
    )


async def _fetch_bulk_report(*, session_id: str | None = None) -> dict[str, Any] | None:
    if session_id:
        doc = await _bulk_reports_collection().find_one({"session_id": session_id})
    else:
        doc = await _bulk_reports_collection().find_one(
            {},
            sort=[("validated_at", -1), ("created_at", -1)],
        )

    if not doc:
        return None

    doc.pop("_id", None)
    return doc


def _build_shopify_links(shopify_id: object) -> dict | None:
    if not shopify_id:
        return None
    try:
        pid = int(shopify_id)
    except Exception:
        return None

    if settings.SHOPIFY_STORE_URL_PROD:
        store = str(settings.SHOPIFY_STORE_URL_PROD).strip()
        if store.startswith("http://") or store.startswith("https://"):
            base = store.rstrip("/")
        else:
            base = f"https://{store}".rstrip("/")
        return {"prod": f"{base}/admin/products/{pid}"}
    return None


def _find_latest_excel_review_path() -> Path:
    root = Path(".")
    candidates: list[Path] = []
    for pattern in EXCEL_REVIEW_PATTERNS:
        candidates.extend(root.glob(pattern))

    if not candidates:
        raise HTTPException(
            status_code=404,
            detail="No Etsy review workbook found. Generate the Excel report first.",
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _load_excel_rows(workbook_path: Path) -> tuple[list[dict], dict[str, int]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], {}

    header_map = {_normalize_header(h): i for i, h in enumerate(headers)}
    out_rows: list[dict] = []
    for values in rows_iter:
        row: dict = {}
        for key, idx in header_map.items():
            row[key] = values[idx] if idx < len(values) else None
        out_rows.append(row)
    return out_rows, header_map


def _extract_etsy_image_from_doc(etsy_doc: dict | None) -> str | None:
    if not etsy_doc:
        return None

    raw = (etsy_doc.get("raw") or {}) if isinstance(etsy_doc, dict) else {}
    images = raw.get("images")
    if isinstance(images, list) and images:
        image = images[0] if isinstance(images[0], dict) else None
        if image:
            for key in ("url_fullxfull", "url_570xN", "url_170x135", "url_75x75"):
                val = image.get(key)
                if val:
                    return str(val)
    return None


async def _resolve_etsy_auth_headers_for_review() -> dict[str, str]:
    try:
        token = await get_valid_etsy_token()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if settings.ETSY_CLIENT_ID and settings.ETSY_CLIENT_SECRET:
        api_key = f"{settings.ETSY_CLIENT_ID}:{settings.ETSY_CLIENT_SECRET}"
    else:
        api_key = settings.ETSY_CLIENT_ID
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing Etsy API key")

    return {
        "Authorization": f"Bearer {token}",
        "x-api-key": str(api_key),
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def _money_to_float(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        amount = value.get("amount")
        divisor = value.get("divisor")
        if amount is None or divisor in (None, 0):
            raise ValueError(f"invalid_money_value={value!r}")
        return float(amount) / float(divisor)
    raise ValueError(f"unsupported_price_value={value!r}")


def _build_etsy_inventory_update_payload(inventory: dict, new_sku: str, target_quantity: int) -> dict:
    products = inventory.get("products") or []
    if not products:
        raise HTTPException(status_code=400, detail="etsy_inventory_missing_products")

    payload_products: list[dict] = []
    for product in products:
        offerings = product.get("offerings") or []
        if not offerings:
            raise HTTPException(status_code=400, detail="etsy_inventory_product_missing_offerings")

        payload_offerings: list[dict] = []
        positive_quantity_found = False
        for offering in offerings:
            quantity = int(offering.get("quantity") or 0)
            if quantity > 0:
                positive_quantity_found = True
            payload_offerings.append(
                {
                    "price": _money_to_float(offering.get("price")),
                    "quantity": quantity,
                    "is_enabled": bool(offering.get("is_enabled")),
                    "readiness_state_id": offering.get("readiness_state_id"),
                }
            )

        if not positive_quantity_found and target_quantity > 0:
            first_enabled_index = next(
                (index for index, item in enumerate(payload_offerings) if item["is_enabled"]),
                0,
            )
            payload_offerings[first_enabled_index]["quantity"] = int(target_quantity)

        payload_product: dict = {
            "sku": str(new_sku),
            "offerings": payload_offerings,
        }
        property_values = product.get("property_values") or []
        if property_values:
            payload_product["property_values"] = copy.deepcopy(property_values)
        payload_products.append(payload_product)

    return {
        "products": payload_products,
        "price_on_property": copy.deepcopy(inventory.get("price_on_property") or []),
        "quantity_on_property": copy.deepcopy(inventory.get("quantity_on_property") or []),
        "sku_on_property": copy.deepcopy(inventory.get("sku_on_property") or []),
        "readiness_state_on_property": copy.deepcopy(inventory.get("readiness_state_on_property") or []),
    }


async def _update_etsy_listing_sku_for_review(
    *,
    listing_id: int,
    new_sku: str,
    target_quantity: int,
    headers: dict[str, str],
) -> None:
    async with httpx.AsyncClient(timeout=45.0) as client:
        get_response = await client.get(
            f"{ETSY_BASE_URL}/listings/{listing_id}/inventory",
            headers=headers,
            params={"show_deleted": "true"},
        )
        if get_response.status_code >= 400:
            raise HTTPException(
                status_code=400,
                detail=f"etsy_inventory_get_failed status={get_response.status_code}",
            )

        inventory_payload = get_response.json() if get_response.content else {}
        update_payload = _build_etsy_inventory_update_payload(inventory_payload, new_sku, target_quantity)

        put_response = await client.put(
            f"{ETSY_BASE_URL}/listings/{listing_id}/inventory",
            headers=headers,
            content=json.dumps(update_payload),
        )
        if put_response.status_code >= 400:
            raise HTTPException(
                status_code=400,
                detail=f"etsy_inventory_put_failed status={put_response.status_code}",
            )


async def _fetch_etsy_main_image_from_api(
    client: httpx.AsyncClient,
    headers: dict[str, str],
    listing_id: int,
) -> str | None:
    response = await client.get(
        f"{ETSY_BASE_URL}/listings/{listing_id}/images",
        headers=headers,
    )
    if response.status_code >= 400:
        return None
    payload = response.json() if response.content else {}
    results = payload.get("results") or []
    if not results:
        return None

    first = results[0] if isinstance(results[0], dict) else None
    if not first:
        return None
    for key in ("url_fullxfull", "url_570xN", "url_170x135", "url_75x75"):
        value = first.get(key)
        if value:
            return str(value)
    return None


def _guess_image_filename(image_url: str, index: int) -> str:
    parsed = urlparse(image_url)
    name = Path(parsed.path or "").name
    if name:
        return name
    return f"image_{index}.jpg"


def _normalize_image_content_type(content_type: str | None, filename: str) -> str:
    txt = (content_type or "").split(";")[0].strip().lower()
    if txt.startswith("image/"):
        return txt
    guessed, _ = mimetypes.guess_type(filename)
    if guessed and guessed.startswith("image/"):
        return guessed
    return "image/jpeg"


async def _upload_etsy_listing_images(
    *,
    shop_id: str | int,
    listing_id: int,
    image_urls: list[str],
    headers: dict[str, str],
) -> dict[str, Any]:
    if not image_urls:
        return {
            "attempted": False,
            "ok": True,
            "uploaded": 0,
            "failed": 0,
            "results": [],
        }

    upload_url = f"{ETSY_BASE_URL}/shops/{shop_id}/listings/{listing_id}/images"
    results: list[dict[str, Any]] = []
    uploaded = 0
    failed = 0

    # Keep auth headers but let httpx set multipart Content-Type with boundary.
    upload_headers = {k: v for k, v in headers.items() if k.lower() != "content-type"}

    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        for idx, image_url in enumerate(image_urls, start=1):
            entry: dict[str, Any] = {
                "rank": idx,
                "image_url": image_url,
                "uploaded": False,
            }
            try:
                image_response = await client.get(image_url)
                if image_response.status_code >= 300:
                    failed += 1
                    entry["error"] = "image_download_failed"
                    entry["download_status"] = image_response.status_code
                    results.append(entry)
                    continue

                filename = _guess_image_filename(image_url, idx)
                content_type = _normalize_image_content_type(image_response.headers.get("content-type"), filename)
                files = {"image": (filename, image_response.content, content_type)}
                data = {"rank": str(idx)}

                upload_response = await client.post(upload_url, headers=upload_headers, files=files, data=data)
                entry["upload_status"] = upload_response.status_code
                if upload_response.status_code < 300:
                    entry["uploaded"] = True
                    uploaded += 1
                    try:
                        entry["etsy_response"] = upload_response.json()
                    except Exception:
                        entry["etsy_response"] = upload_response.text
                else:
                    failed += 1
                    entry["error"] = "etsy_image_upload_failed"
                    try:
                        entry["etsy_response"] = upload_response.json()
                    except Exception:
                        entry["etsy_response"] = upload_response.text
            except Exception as exc:
                failed += 1
                entry["error"] = "image_upload_exception"
                entry["message"] = str(exc)

            results.append(entry)

    return {
        "attempted": True,
        "ok": failed == 0,
        "uploaded": uploaded,
        "failed": failed,
        "results": results,
    }


def _remove_listing_from_excel(workbook_path: Path, listing_id: int) -> bool:
    wb = load_workbook(workbook_path)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    if "listing_id" not in headers:
        raise HTTPException(status_code=400, detail="Workbook missing listing_id column")

    listing_col = headers.index("listing_id") + 1
    target_row = None
    for row_num in range(2, ws.max_row + 1):
        value = ws.cell(row=row_num, column=listing_col).value
        try:
            if int(value) == int(listing_id):
                target_row = row_num
                break
        except Exception:
            continue

    if target_row is None:
        wb.close()
        return False

    ws.delete_rows(target_row, 1)
    wb.save(workbook_path)
    wb.close()
    return True


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, dict):
        amount = value.get("amount")
        divisor = value.get("divisor")
        if amount is None:
            return None
        try:
            amount_f = float(amount)
            divisor_f = float(divisor or 1)
            if divisor_f == 0:
                return None
            return round(amount_f / divisor_f, 2)
        except Exception:
            return None
    try:
        return float(value)
    except Exception:
        return None


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except Exception:
        return None


def _cmp_values(left: object, right: object, *, numeric: bool = False) -> str:
    if left is None and right is None:
        return "missing"
    if left is None or right is None:
        return "missing"

    if numeric:
        left_f = _to_float(left)
        right_f = _to_float(right)
        if left_f is None or right_f is None:
            return "diff"
        return "same" if abs(left_f - right_f) < 0.01 else "diff"

    return "same" if str(left).strip() == str(right).strip() else "diff"


def _strip_html(text: str) -> str:
    return re.sub(r"<[^>]+>", " ", text)


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    txt = _strip_html(txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt or None


def _clean_list(value: object, *, limit: int | None = None) -> list[str]:
    if not isinstance(value, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for raw in value:
        txt = _clean_text(raw)
        if not txt:
            continue
        key = txt.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(txt)
        if limit is not None and len(out) >= limit:
            break
    return out


def _first_attr(attrs: dict, keys: list[str]) -> object:
    if not isinstance(attrs, dict):
        return None
    lowered = {str(k).strip().lower(): v for k, v in attrs.items()}
    for key in keys:
        if key in lowered:
            return lowered[key]
    return None


def _extract_materials(attrs: dict, fallback_tags: list[str]) -> list[str]:
    raw = _first_attr(attrs, ["material", "materials", "primary material", "primary materials"])
    if isinstance(raw, list):
        materials = _clean_list(raw, limit=13)
    elif raw is None:
        materials = []
    else:
        parts = re.split(r"[,/;|]", str(raw))
        materials = _clean_list(parts, limit=13)

    if materials:
        return materials[:13]

    candidates = [t for t in fallback_tags if len(t) <= 20]
    return candidates[:5]


def _coerce_etsy_enum(value: object, allowed: set[str]) -> str | None:
    txt = _clean_text(value)
    if not txt:
        return None
    txt = txt.lower().replace("-", "_").replace(" ", "_")
    return txt if txt in allowed else None


def _guess_who_made(attrs: dict) -> str | None:
    direct = _coerce_etsy_enum(_first_attr(attrs, ["who made", "who_made", "maker"]), ETSY_ALLOWED_WHO_MADE)
    if direct:
        return direct

    handmade = _clean_text(_first_attr(attrs, ["handmade", "is handmade", "is_handmade"]))
    if handmade and handmade.lower() in {"yes", "true", "1"}:
        return "i_did"
    return None


def _guess_when_made(attrs: dict) -> str | None:
    direct = _coerce_etsy_enum(_first_attr(attrs, ["when made", "when_made"]), ETSY_ALLOWED_WHEN_MADE)
    if direct:
        return direct

    year_raw = _clean_text(_first_attr(attrs, ["year made", "year_made", "year", "era"]))
    if not year_raw:
        return None
    if re.search(r"\bbefore\s+1700\b", year_raw.lower()):
        return "before_1700"
    m = re.search(r"(1[7-9]\d{2}|20\d{2})", year_raw)
    if not m:
        # Handles values like "19th century" when exact year is not present.
        c = re.search(r"\b(1[7-9]|20)(st|nd|rd|th)\s*century\b", year_raw.lower())
        if not c:
            return None
        century_number = int(c.group(1))
        year = (century_number - 1) * 100
    else:
        year = int(m.group(1))

    if year >= 2020:
        return "2020_2026"
    if year >= 2010:
        return "2010_2019"
    if year >= 2007:
        return "2007_2009"
    if year >= 2000:
        return "2000_2006"
    if year >= 1990:
        return "1990s"
    if year >= 1980:
        return "1980s"
    if year >= 1970:
        return "1970s"
    if year >= 1960:
        return "1960s"
    if year >= 1950:
        return "1950s"
    if year >= 1940:
        return "1940s"
    if year >= 1930:
        return "1930s"
    if year >= 1920:
        return "1920s"
    if year >= 1910:
        return "1910s"
    if year >= 1900:
        return "1900s"
    if year >= 1800:
        return "1800s"
    if year >= 1700:
        return "1700s"
    return "before_1700"


def _is_missing_etsy_required(key: str, value: object, *, listing_type: str | None = None) -> bool:
    if key in {"title", "description", "who_made", "when_made"}:
        txt = _clean_text(value)
        if not txt:
            return True
        if key == "who_made":
            return txt not in ETSY_ALLOWED_WHO_MADE
        if key == "when_made":
            return txt not in ETSY_ALLOWED_WHEN_MADE
        return False

    if key == "type":
        txt = _clean_text(value)
        return not txt or txt not in ETSY_ALLOWED_LISTING_TYPES

    if key == "price":
        v = _to_float(value)
        return v is None or v <= 0
    if key == "quantity":
        v = _to_int(value)
        return v is None or v <= 0

    # shipping_profile_id only required for physical listings
    if key == "shipping_profile_id":
        if listing_type == "digital":
            return False
        v = _to_int(value)
        return v is None or v <= 0

    if key in {"taxonomy_id", "return_policy_id"}:
        v = _to_int(value)
        return v is None or v <= 0

    return value is None


def _fmt_measure(m: dict | None) -> str | None:
    if not isinstance(m, dict):
        return None
    val = m.get("value") or m.get("amount")
    unit = m.get("unit")
    if val is None:
        return None
    return f"{val} {unit}" if unit else str(val)


def _fmt_weight(weight: dict | None) -> str | None:
    if not isinstance(weight, dict):
        return None
    major = _fmt_measure(weight.get("major"))
    minor = _fmt_measure(weight.get("minor"))
    parts = [p for p in [major, minor] if p]
    return " ".join(parts) if parts else None


def _fmt_dimensions(dims: dict | None) -> str | None:
    if not isinstance(dims, dict):
        return None
    parts = []
    for key in ("length", "width", "height"):
        v = _fmt_measure(dims.get(key))
        if v:
            parts.append(f"{key[0].upper()}: {v}")
    return ", ".join(parts) if parts else None


def _normalize_etsy_weight_unit(unit: object) -> str | None:
    raw = (_clean_text(unit) or "").lower()
    mapping = {
        "lb": "lb",
        "lbs": "lb",
        "pound": "lb",
        "pounds": "lb",
        "oz": "oz",
        "ounce": "oz",
        "ounces": "oz",
        "g": "g",
        "gram": "g",
        "grams": "g",
        "kg": "kg",
        "kilogram": "kg",
        "kilograms": "kg",
    }
    return mapping.get(raw)


def _normalize_etsy_dimension_unit(unit: object) -> str | None:
    raw = (_clean_text(unit) or "").lower()
    mapping = {
        "in": "inches",
        "inch": "inches",
        "inches": "inches",
        "ft": "ft",
        "foot": "ft",
        "feet": "ft",
        "mm": "mm",
        "millimeter": "mm",
        "millimeters": "mm",
        "cm": "cm",
        "centimeter": "cm",
        "centimeters": "cm",
        "m": "m",
        "meter": "m",
        "meters": "m",
        "yd": "yd",
        "yard": "yd",
        "yards": "yd",
    }
    return mapping.get(raw)


def _coerce_etsy_whole_weight(value: object) -> int | None:
    num = _to_float(value)
    if num is None:
        return None
    if num <= 0:
        return None
    return max(1, int(math.ceil(num)))


def _extract_etsy_shipping_measurements(package: dict | None) -> dict[str, Any]:
    if not isinstance(package, dict):
        return {
            "item_weight": None,
            "item_weight_unit": None,
            "item_length": None,
            "item_width": None,
            "item_height": None,
            "item_dimensions_unit": None,
        }

    weight = package.get("weight") or {}
    major = weight.get("major") or {}
    minor = weight.get("minor") or {}
    major_value = _to_float(major.get("value"))
    minor_value = _to_float(minor.get("value"))
    major_unit = _normalize_etsy_weight_unit(major.get("unit"))
    minor_unit = _normalize_etsy_weight_unit(minor.get("unit"))

    item_weight = None
    item_weight_unit = None
    if major_value is not None and major_unit:
        item_weight = major_value
        item_weight_unit = major_unit
        if major_unit == "lb" and minor_value is not None and minor_unit == "oz":
            item_weight = round(major_value + (minor_value / 16.0), 3)
        elif major_unit == "kg" and minor_value is not None and minor_unit == "g":
            item_weight = round(major_value + (minor_value / 1000.0), 3)
    elif minor_value is not None and minor_unit:
        item_weight = minor_value
        item_weight_unit = minor_unit

    dimensions = package.get("dimensions") or {}
    length = dimensions.get("length") or {}
    width = dimensions.get("width") or {}
    height = dimensions.get("height") or {}
    item_dimensions_unit = (
        _normalize_etsy_dimension_unit(length.get("unit"))
        or _normalize_etsy_dimension_unit(width.get("unit"))
        or _normalize_etsy_dimension_unit(height.get("unit"))
    )

    return {
        "item_weight": _coerce_etsy_whole_weight(item_weight),
        "item_weight_unit": item_weight_unit,
        "item_length": _to_float(length.get("value")),
        "item_width": _to_float(width.get("value")),
        "item_height": _to_float(height.get("value")),
        "item_dimensions_unit": item_dimensions_unit,
    }


def _resolve_etsy_api_key() -> str | None:
    if settings.ETSY_CLIENT_ID and settings.ETSY_CLIENT_SECRET:
        return f"{settings.ETSY_CLIENT_ID}:{settings.ETSY_CLIENT_SECRET}"
    return settings.ETSY_CLIENT_ID


def _tokenize_taxonomy_text(value: object) -> list[str]:
    text = (_clean_text(value) or "").lower()
    if not text:
        return []
    return [token for token in re.split(r"[^a-z0-9]+", text) if len(token) >= 2]


def _dedupe_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value not in seen:
            out.append(value)
            seen.add(value)
    return out


def _resolve_default_etsy_when_made() -> str:
    configured = _coerce_etsy_enum(settings.ETSY_DEFAULT_WHEN_MADE, ETSY_ALLOWED_WHEN_MADE)
    return configured or "before_2007"


def _sanitize_etsy_tag(value: object) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    text = ETSY_ALLOWED_TAG_CHARS_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip(" -'")
    if not text or len(text) > ETSY_TAG_MAX_LENGTH:
        return None
    return text


def _sanitize_etsy_tags(values: list[object], *, limit: int = ETSY_TAG_MAX_COUNT) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        tag = _sanitize_etsy_tag(value)
        if not tag:
            continue
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(tag)
        if len(out) >= limit:
            break
    return out


def _sanitize_etsy_title(value: object) -> str | None:
    txt = _clean_text(value)
    if not txt:
        return None
    txt = re.sub(r"\s+", " ", txt).strip()
    if not txt:
        return None
    return txt[:ETSY_TITLE_MAX_LENGTH]


def _build_etsy_seo_input(detail: dict) -> dict[str, Any]:
    current = detail.get("current") or {}
    optimized = detail.get("etsy_optimized") or {}
    return {
        "title": _clean_text(detail.get("title")) or "",
        "description": _clean_text(current.get("description") or optimized.get("description")) or "",
        "category": _clean_text(detail.get("category")) or "",
        "tags": _sanitize_etsy_tags(current.get("tags") or optimized.get("tags") or []),
        "materials": _clean_list(optimized.get("materials") or [], limit=6),
    }


def _suggest_etsy_seo_with_openai(seo_input: dict[str, Any]) -> dict[str, Any] | None:
    if not settings.OPENAI_API_KEY:
        return None

    client = OpenAI(api_key=settings.OPENAI_API_KEY)
    payload = {
        "product": seo_input,
        "when_made_allowed": sorted(ETSY_ALLOWED_WHEN_MADE),
        "rules": [
            "Return one Etsy SEO-focused title that is at most 140 characters.",
            "The title must be readable and include strong search intent keywords.",
            "Infer when_made from title, tags and description, selecting one allowed enum.",
            "If evidence is weak, prefer a conservative vintage bucket from allowed values.",
            "Do not invent facts not present in product context.",
        ],
    }
    response = client.responses.create(
        model=ETSY_SEO_OPENAI_MODEL,
        input=[
            {
                "role": "system",
                "content": "You are an Etsy and SEO expert. Produce compliant Etsy listing title suggestions and infer when_made safely.",
            },
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "etsy_seo_optimization",
                "schema": {
                    "type": "object",
                    "properties": {
                        "seo_title": {"type": "string"},
                        "when_made": {"type": "string"},
                        "reason": {"type": "string"},
                    },
                    "required": ["seo_title", "when_made", "reason"],
                    "additionalProperties": False,
                },
            }
        },
    )

    try:
        parsed = json.loads((response.output_text or "").strip())
    except Exception:
        return None

    seo_title = _sanitize_etsy_title(parsed.get("seo_title"))
    when_made = _coerce_etsy_enum(parsed.get("when_made"), ETSY_ALLOWED_WHEN_MADE) or _resolve_default_etsy_when_made()
    if not seo_title:
        return None

    return {
        "seo_title": seo_title,
        "when_made": when_made,
        "reason": _clean_text(parsed.get("reason")) or "",
    }


async def _generate_etsy_seo_optimization(detail: dict) -> dict[str, Any]:
    seo_input = _build_etsy_seo_input(detail)
    
    # Rate limit OpenAI requests
    semaphore = _get_openai_semaphore()
    async with semaphore:
        ai_result = await asyncio.to_thread(_suggest_etsy_seo_with_openai, seo_input)
        await asyncio.sleep(OPENAI_REQUEST_DELAY_SECONDS)

    fallback_title = _sanitize_etsy_title((detail.get("etsy_optimized") or {}).get("title") or detail.get("title"))
    fallback_when_made = (
        _coerce_etsy_enum((detail.get("etsy_optimized") or {}).get("when_made"), ETSY_ALLOWED_WHEN_MADE)
        or _coerce_etsy_enum((detail.get("current") or {}).get("when_made"), ETSY_ALLOWED_WHEN_MADE)
        or _resolve_default_etsy_when_made()
    )

    return {
        "ai_used": bool(ai_result),
        "reason": (ai_result or {}).get("reason"),
        "seo_title": (ai_result or {}).get("seo_title") or fallback_title,
        "when_made": (ai_result or {}).get("when_made") or fallback_when_made,
    }


def _build_etsy_tag_suggestion_input(detail: dict) -> dict[str, Any]:
    current = detail.get("current") or {}
    optimized = detail.get("etsy_optimized") or {}
    return {
        "title": _clean_text(optimized.get("title") or detail.get("title")) or "",
        "category": _clean_text(detail.get("category")) or "",
        "description": _clean_text(optimized.get("description")) or "",
        "current_tags": _sanitize_etsy_tags(current.get("tags") or []),
        "materials": _clean_list(optimized.get("materials") or [], limit=6),
    }


def _suggest_etsy_tags_with_openai(tag_input: dict[str, Any]) -> dict[str, Any] | None:
    if not settings.OPENAI_API_KEY:
        return None

    client = OpenAI(api_key=settings.OPENAI_API_KEY)
    payload = {
        "product": tag_input,
        "rules": [
            "Return up to 13 Etsy tags.",
            "Each tag must be 20 characters or fewer.",
            "Use only letters, numbers, spaces, hyphen, apostrophe, and the symbols ™ © ®.",
            "Prefer short Etsy search phrases over long descriptions.",
            "Do not repeat tags or invent facts that are not supported by the product data.",
        ],
    }
    response = client.responses.create(
        model=ETSY_TAG_OPENAI_MODEL,
        input=[
            {
                "role": "system",
                "content": "You write Etsy-safe product tags. Follow the character and length limits exactly and return only valid JSON.",
            },
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "etsy_tag_suggestions",
                "schema": {
                    "type": "object",
                    "properties": {
                        "tags": {
                            "type": "array",
                            "items": {"type": "string"},
                            "maxItems": ETSY_TAG_MAX_COUNT,
                        },
                        "reason": {"type": "string"},
                    },
                    "required": ["tags", "reason"],
                    "additionalProperties": False,
                },
            }
        },
    )

    try:
        parsed = json.loads((response.output_text or "").strip())
    except Exception:
        return None

    tags = _sanitize_etsy_tags(parsed.get("tags") or [])
    if not tags:
        return None

    return {
        "tags": tags,
        "reason": _clean_text(parsed.get("reason")) or "",
    }


async def _generate_etsy_tag_suggestions(detail: dict) -> dict[str, Any]:
    tag_input = _build_etsy_tag_suggestion_input(detail)
    fallback_tags = tag_input.get("current_tags") or []
    
    # Rate limit OpenAI requests
    semaphore = _get_openai_semaphore()
    async with semaphore:
        ai_result = await asyncio.to_thread(_suggest_etsy_tags_with_openai, tag_input)
        await asyncio.sleep(OPENAI_REQUEST_DELAY_SECONDS)
    
    tags = (ai_result or {}).get("tags") or fallback_tags
    return {
        "ai_used": bool(ai_result),
        "reason": (ai_result or {}).get("reason"),
        "tags": tags[:ETSY_TAG_MAX_COUNT],
        "source_tags": fallback_tags,
    }


def _build_etsy_taxonomy_query(detail: dict) -> dict[str, Any]:
    optimized = detail.get("etsy_optimized") or {}
    title = _clean_text(optimized.get("title") or detail.get("title")) or ""
    category = _clean_text(detail.get("category")) or ""
    tags = _clean_list(optimized.get("tags") or [])
    materials = _clean_list(optimized.get("materials") or [])
    query_parts = [title, category, " ".join(tags[:8]), " ".join(materials[:6])]
    query_text = " | ".join(part for part in query_parts if part)
    query_tokens = _dedupe_keep_order(
        _tokenize_taxonomy_text(title)
        + _tokenize_taxonomy_text(category)
        + [token for tag in tags[:8] for token in _tokenize_taxonomy_text(tag)]
        + [token for material in materials[:6] for token in _tokenize_taxonomy_text(material)]
    )
    return {
        "title": title,
        "category": category,
        "tags": tags,
        "materials": materials,
        "query_text": query_text,
        "query_tokens": query_tokens,
    }


async def _fetch_buyer_taxonomy_nodes() -> list[dict[str, Any]]:
    cached_nodes = _ETSY_BUYER_TAXONOMY_CACHE.get("nodes")
    fetched_at = _ETSY_BUYER_TAXONOMY_CACHE.get("fetched_at")
    now = _now_utc()
    if cached_nodes and isinstance(fetched_at, datetime):
        age = (now - fetched_at).total_seconds()
        if age < ETSY_TAXONOMY_CACHE_TTL_SECONDS:
            return cached_nodes

    api_key = _resolve_etsy_api_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing Etsy API key")

    headers = {
        "x-api-key": str(api_key),
        "Accept": "application/json",
    }
    url = f"{ETSY_BASE_URL}/buyer-taxonomy/nodes"

    async with httpx.AsyncClient(timeout=45.0) as client:
        response = await client.get(url, headers=headers)

    if response.status_code >= 300:
        try:
            detail = response.json()
        except Exception:
            detail = response.text
        raise HTTPException(
            status_code=502,
            detail={
                "error": "etsy_taxonomy_fetch_failed",
                "etsy_status_code": response.status_code,
                "etsy_response": detail,
            },
        )

    payload = response.json() if response.text else {}
    nodes = payload.get("results") or []
    _ETSY_BUYER_TAXONOMY_CACHE["nodes"] = nodes
    _ETSY_BUYER_TAXONOMY_CACHE["fetched_at"] = now
    return nodes


def _flatten_buyer_taxonomy_nodes(nodes: list[dict[str, Any]], ancestors: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    ancestors = ancestors or []
    flat: list[dict[str, Any]] = []
    for node in nodes or []:
        name = _clean_text(node.get("name"))
        taxonomy_id = _to_int(node.get("id"))
        if not name or taxonomy_id is None:
            continue

        lineage = [*ancestors, {"id": taxonomy_id, "name": name}]
        full_path_names = [item["name"] for item in lineage]
        full_path_ids = [item["id"] for item in lineage]
        children = node.get("children") or []
        leaf = not bool(children)
        full_path = " > ".join(full_path_names)
        flat.append(
            {
                "taxonomy_id": taxonomy_id,
                "name": name,
                "full_path": full_path,
                "full_path_taxonomy_ids": full_path_ids,
                "level": _to_int(node.get("level")) or max(0, len(lineage) - 1),
                "leaf": leaf,
                "name_tokens": _tokenize_taxonomy_text(name),
                "path_tokens": _tokenize_taxonomy_text(full_path),
            }
        )
        if children:
            flat.extend(_flatten_buyer_taxonomy_nodes(children, lineage))
    return flat


def _score_etsy_taxonomy_candidate(query: dict[str, Any], candidate: dict[str, Any]) -> float:
    tokens = set(query.get("query_tokens") or [])
    if not tokens:
        return 0.0

    name_tokens = set(candidate.get("name_tokens") or [])
    path_tokens = set(candidate.get("path_tokens") or [])
    name_overlap = tokens & name_tokens
    path_overlap = tokens & path_tokens

    score = 0.0
    score += len(name_overlap) * 4.0
    score += len(path_overlap) * 1.5

    title = (query.get("title") or "").lower()
    category = (query.get("category") or "").lower()
    name = (candidate.get("name") or "").lower()
    full_path = (candidate.get("full_path") or "").lower()

    if name and name in title:
        score += 6.0
    if name and name in category:
        score += 3.0
    if full_path and category and category in full_path:
        score += 2.0

    if candidate.get("leaf"):
        score += 1.0

    level = candidate.get("level") or 0
    if level <= 1:
        score -= 2.0

    return round(score, 3)


def _prefilter_etsy_taxonomy_candidates(query: dict[str, Any], flat_nodes: list[dict[str, Any]], *, limit: int = ETSY_TAXONOMY_PREFILTER_LIMIT) -> list[dict[str, Any]]:
    scored: list[dict[str, Any]] = []
    for node in flat_nodes:
        score = _score_etsy_taxonomy_candidate(query, node)
        if score <= 0:
            continue
        scored.append({
            "taxonomy_id": node["taxonomy_id"],
            "name": node["name"],
            "full_path": node["full_path"],
            "level": node["level"],
            "leaf": bool(node.get("leaf")),
            "local_score": score,
        })

    scored.sort(key=lambda item: (item["local_score"], item["leaf"], item["level"]), reverse=True)
    return scored[:limit]


def _rank_etsy_taxonomy_candidates_with_openai(query: dict[str, Any], candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not settings.OPENAI_API_KEY or not candidates:
        return None

    client = OpenAI(api_key=settings.OPENAI_API_KEY)
    candidate_ids = {item["taxonomy_id"] for item in candidates}
    payload = {
        "product": {
            "title": query.get("title"),
            "category": query.get("category"),
            "tags": query.get("tags") or [],
            "materials": query.get("materials") or [],
            "query_text": query.get("query_text"),
        },
        "candidates": candidates,
        "rules": [
            "Choose exactly one taxonomy_id from the candidate list when there is a reasonable fit.",
            "Prefer the most specific, leaf-level category that fits the product.",
            "Do not invent taxonomy ids.",
            "If the fit is weak, return null and low confidence.",
        ],
    }
    system_message = (
        "You choose Etsy taxonomy categories for product listings. "
        "Only select from provided candidates. Prefer the most specific valid category."
    )
    response = client.responses.create(
        model=ETSY_TAXONOMY_OPENAI_MODEL,
        input=[
            {"role": "system", "content": system_message},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "etsy_taxonomy_choice",
                "schema": {
                    "type": "object",
                    "properties": {
                        "taxonomy_id": {"type": ["integer", "null"]},
                        "confidence": {"type": "number"},
                        "reason": {"type": "string"}
                    },
                    "required": ["taxonomy_id", "confidence", "reason"],
                    "additionalProperties": False
                }
            }
        },
    )

    try:
        parsed = json.loads((response.output_text or "").strip())
    except Exception:
        return None

    taxonomy_id = _to_int(parsed.get("taxonomy_id"))
    if taxonomy_id is not None and taxonomy_id not in candidate_ids:
        taxonomy_id = None

    confidence = parsed.get("confidence")
    try:
        confidence_value = max(0.0, min(1.0, float(confidence)))
    except Exception:
        confidence_value = 0.0

    return {
        "taxonomy_id": taxonomy_id,
        "confidence": round(confidence_value, 3),
        "reason": _clean_text(parsed.get("reason")) or "",
    }


async def _suggest_etsy_taxonomy(detail: dict, *, limit: int = ETSY_TAXONOMY_RESULT_LIMIT) -> dict[str, Any]:
    query = _build_etsy_taxonomy_query(detail)
    raw_nodes = await _fetch_buyer_taxonomy_nodes()
    flat_nodes = _flatten_buyer_taxonomy_nodes(raw_nodes)
    prefiltered = _prefilter_etsy_taxonomy_candidates(query, flat_nodes)

    # Rate limit OpenAI requests
    semaphore = _get_openai_semaphore()
    async with semaphore:
        ai_choice = await asyncio.to_thread(_rank_etsy_taxonomy_candidates_with_openai, query, prefiltered)
        await asyncio.sleep(OPENAI_REQUEST_DELAY_SECONDS)
    
    suggestions = [dict(item) for item in prefiltered[:limit]]
    best_match = None

    if ai_choice and ai_choice.get("taxonomy_id") is not None:
        chosen_id = ai_choice["taxonomy_id"]
        suggestions.sort(key=lambda item: (item["taxonomy_id"] == chosen_id, item["local_score"]), reverse=True)
        for item in suggestions:
            if item["taxonomy_id"] == chosen_id:
                item["ai_selected"] = True
                item["ai_confidence"] = ai_choice.get("confidence")
                item["ai_reason"] = ai_choice.get("reason")
                best_match = item
                break

    if best_match is None and suggestions:
        best_match = suggestions[0]

    return {
        "query": query,
        "ai_used": bool(ai_choice),
        "ai_choice": ai_choice,
        "best_match": best_match,
        "suggestions": suggestions,
    }


async def _resolve_etsy_shop_id() -> str | None:
    doc = await db.product_normalized.find_one(
        {"channels.etsy.shop_id": {"$exists": True, "$ne": None}},
        {"channels.etsy.shop_id": 1},
    )
    shop_id = ((doc or {}).get("channels") or {}).get("etsy", {}).get("shop_id")
    if shop_id:
        return str(shop_id)

    etsy_doc = await db.etsy_listings_investigation.find_one(
        {"shop_id": {"$exists": True, "$ne": None}},
        {"shop_id": 1},
    )
    if etsy_doc and etsy_doc.get("shop_id") is not None:
        return str(etsy_doc.get("shop_id"))
    return None


async def _fetch_etsy_readiness_states(shop_id: str) -> list[dict[str, Any]]:
    cached_by_shop = _ETSY_READINESS_CACHE.setdefault("by_shop", {})
    cached = cached_by_shop.get(str(shop_id))
    now = _now_utc()
    if isinstance(cached, dict) and isinstance(cached.get("fetched_at"), datetime):
        age = (now - cached["fetched_at"]).total_seconds()
        if age < ETSY_TAXONOMY_CACHE_TTL_SECONDS:
            return cached.get("results") or []

    token = await get_valid_etsy_token()
    api_key = _resolve_etsy_api_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing Etsy API key")

    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": str(api_key),
        "Accept": "application/json",
    }
    url = f"{ETSY_BASE_URL}/shops/{shop_id}/readiness-state-definitions"

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(url, headers=headers)

    if response.status_code >= 300:
        try:
            detail = response.json()
        except Exception:
            detail = response.text
        raise HTTPException(
            status_code=502,
            detail={
                "error": "etsy_readiness_state_fetch_failed",
                "etsy_status_code": response.status_code,
                "etsy_response": detail,
            },
        )

    payload = response.json() if response.text else {}
    results = payload.get("results") or []
    cached_by_shop[str(shop_id)] = {"fetched_at": now, "results": results}
    return results


async def _resolve_etsy_readiness_state_id(*, shop_id: str, payload: dict[str, Any]) -> int | None:
    configured = _to_int(payload.get("readiness_state_id")) or settings.ETSY_READINESS_STATE_ID
    if configured:
        return int(configured)

    if payload.get("type") != "physical":
        return None

    readiness_states = await _fetch_etsy_readiness_states(shop_id)
    if not readiness_states:
        return None

    preferred_state = "made_to_order" if payload.get("when_made") == "made_to_order" else "ready_to_ship"
    fallback_id = None
    for item in readiness_states:
        readiness_state_id = _to_int(item.get("readiness_state_id"))
        if readiness_state_id is None:
            continue
        if fallback_id is None:
            fallback_id = readiness_state_id
        if _clean_text(item.get("readiness_state")) == preferred_state:
            return readiness_state_id
    return fallback_id


async def _record_etsy_draft_attempt(
    *,
    sku: str,
    shop_id: str | None,
    request_payload: dict[str, Any] | None,
    request_form_data: dict[str, Any] | None,
    response_status_code: int | None,
    response_payload: Any,
    ok: bool,
    linked_listing_id: int | None = None,
    error_code: str | None = None,
) -> None:
    attempted_at = _now_utc()
    attempt_doc = {
        "sku": sku,
        "shop_id": shop_id,
        "ok": bool(ok),
        "error_code": error_code,
        "etsy_status_code": response_status_code,
        "request_payload": request_payload,
        "request_form_data": request_form_data,
        "response": response_payload,
        "linked_listing_id": linked_listing_id,
        "attempted_at": attempted_at,
    }
    await db.etsy_create_draft_attempts.insert_one(attempt_doc)

    set_fields: dict[str, Any] = {
        "channels.etsy.last_create_draft_request": request_payload,
        "channels.etsy.last_create_draft_response": response_payload,
        "channels.etsy.last_create_draft_http_status": response_status_code,
        "channels.etsy.last_create_draft_ok": bool(ok),
        "channels.etsy.last_create_draft_error_code": error_code,
        "channels.etsy.last_create_draft_at": attempted_at,
    }
    if request_form_data is not None:
        set_fields["channels.etsy.last_create_draft_form_data"] = request_form_data
    if shop_id is not None:
        set_fields["channels.etsy.shop_id"] = _to_int(shop_id) or shop_id
    if linked_listing_id is not None:
        set_fields["channels.etsy.listing_id"] = linked_listing_id
        set_fields["channels.etsy.listing_state"] = "draft"

    await db.product_normalized.update_one(
        {"$or": [{"_id": sku}, {"sku": sku}]},
        {
            "$set": set_fields,
            "$push": {
                "channels.etsy.create_draft_attempts": {
                    "$each": [
                        {
                            "ok": bool(ok),
                            "error_code": error_code,
                            "etsy_status_code": response_status_code,
                            "linked_listing_id": linked_listing_id,
                            "attempted_at": attempted_at,
                        }
                    ],
                    "$slice": -20,
                }
            },
            "$inc": {"channels.etsy.create_draft_attempt_count": 1},
        },
    )


def _build_etsy_create_listing_payload(detail: dict, payload_override: dict[str, Any] | None = None) -> dict[str, Any]:
    optimized = dict((detail.get("etsy_optimized") or {}))
    if payload_override:
        for key, value in payload_override.items():
            optimized[key] = value

    listing_type = _coerce_etsy_enum(optimized.get("type"), ETSY_ALLOWED_LISTING_TYPES) or settings.ETSY_DEFAULT_LISTING_TYPE

    payload: dict[str, Any] = {
        "title": (_clean_text(optimized.get("title")) or "")[:140] or None,
        "description": _clean_text(optimized.get("description")),
        "quantity": _to_int(optimized.get("quantity")),
        "price": _to_float(optimized.get("price")),
        "type": listing_type,
        "who_made": _coerce_etsy_enum(optimized.get("who_made"), ETSY_ALLOWED_WHO_MADE),
        "when_made": _coerce_etsy_enum(optimized.get("when_made"), ETSY_ALLOWED_WHEN_MADE) or _resolve_default_etsy_when_made(),
        "taxonomy_id": _to_int(optimized.get("taxonomy_id")),
        "shipping_profile_id": _to_int(optimized.get("shipping_profile_id")),
        "return_policy_id": _to_int(optimized.get("return_policy_id")),
        "readiness_state_id": _to_int(optimized.get("readiness_state_id")) or settings.ETSY_READINESS_STATE_ID,
        "item_weight": _coerce_etsy_whole_weight(optimized.get("item_weight")),
        "item_weight_unit": _clean_text(optimized.get("item_weight_unit")),
        "item_length": _to_float(optimized.get("item_length")),
        "item_width": _to_float(optimized.get("item_width")),
        "item_height": _to_float(optimized.get("item_height")),
        "item_dimensions_unit": _clean_text(optimized.get("item_dimensions_unit")),
        "tags": _sanitize_etsy_tags(optimized.get("tags") or []),
        "materials": _clean_list(optimized.get("materials") or [], limit=13),
    }

    if payload["quantity"] is not None:
        payload["quantity"] = max(1, int(payload["quantity"]))

    if payload.get("type") == "digital":
        payload["shipping_profile_id"] = None
        payload["readiness_state_id"] = None

    return payload


def _get_missing_etsy_required_fields(payload: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    listing_type = payload.get("type")
    for key, _label in ETSY_REQUIRED_CREATE_FIELDS:
        if _is_missing_etsy_required(key, payload.get(key), listing_type=listing_type):
            missing.append(key)
    return missing


def _to_etsy_form_data(payload: dict[str, Any]) -> dict[str, Any]:
    # Etsy create listing expects form-like fields. Lists are sent as comma-separated values.
    out: dict[str, Any] = {}
    for key, value in payload.items():
        if value is None:
            continue
        if isinstance(value, list):
            if value:
                out[key] = ",".join(str(v) for v in value if v is not None)
            continue
        out[key] = str(value)
    return out


def _build_etsy_publish_comparison(doc: dict) -> dict:
    channels = doc.get("channels") or {}
    etsy = channels.get("etsy") or {}
    attrs = doc.get("attributes") or {}
    package = doc.get("package") or {}
    shipping_measurements = _extract_etsy_shipping_measurements(package)

    tags = _sanitize_etsy_tags(doc.get("tags") or etsy.get("tags") or [])
    materials = _clean_list(etsy.get("materials") or [], limit=13) or _extract_materials(attrs, tags)
    images = _clean_list(doc.get("images") or [], limit=10)

    weight_display = _fmt_weight(package.get("weight"))
    dimensions_display = _fmt_dimensions(package.get("dimensions"))

    listing_type_current = _coerce_etsy_enum(etsy.get("type"), ETSY_ALLOWED_LISTING_TYPES) or settings.ETSY_DEFAULT_LISTING_TYPE

    current = {
        "title": _clean_text(doc.get("title")),
        "description": _clean_text(doc.get("description")),
        "price": _to_float(doc.get("price")),
        "quantity": _to_int(doc.get("quantity")),
        "type": _coerce_etsy_enum(etsy.get("type"), ETSY_ALLOWED_LISTING_TYPES),
        "who_made": _coerce_etsy_enum(etsy.get("who_made"), ETSY_ALLOWED_WHO_MADE),
        "when_made": _coerce_etsy_enum(etsy.get("when_made"), ETSY_ALLOWED_WHEN_MADE),
        "taxonomy_id": _to_int(etsy.get("taxonomy_id")),
        "shipping_profile_id": _to_int(etsy.get("shipping_profile_id")),
        "return_policy_id": _to_int(etsy.get("return_policy_id")),
        "readiness_state_id": _to_int(etsy.get("readiness_state_id")),
        "item_weight": shipping_measurements.get("item_weight"),
        "item_weight_unit": shipping_measurements.get("item_weight_unit"),
        "item_length": shipping_measurements.get("item_length"),
        "item_width": shipping_measurements.get("item_width"),
        "item_height": shipping_measurements.get("item_height"),
        "item_dimensions_unit": shipping_measurements.get("item_dimensions_unit"),
        "tags": tags,
        "materials": _clean_list(etsy.get("materials") or [], limit=13),
        "images": images,
        "weight": weight_display,
        "dimensions": dimensions_display,
    }

    optimized_type = current.get("type") or settings.ETSY_DEFAULT_LISTING_TYPE

    optimized = {
        "title": (_clean_text(doc.get("title")) or "")[:140] or None,
        "description": _clean_text(doc.get("description")),
        "price": _to_float(doc.get("price")),
        "quantity": _to_int(doc.get("quantity")),
        "type": optimized_type,
        "who_made": current.get("who_made") or _guess_who_made(attrs) or settings.ETSY_DEFAULT_WHO_MADE,
        "when_made": current.get("when_made") or _guess_when_made(attrs) or _resolve_default_etsy_when_made(),
        "taxonomy_id": current.get("taxonomy_id"),
        "shipping_profile_id": current.get("shipping_profile_id") or settings.ETSY_SHIPPING_PROFILE_ID,
        "return_policy_id": current.get("return_policy_id") or settings.ETSY_RETURN_POLICY_ID,
        "readiness_state_id": current.get("readiness_state_id") or settings.ETSY_READINESS_STATE_ID,
        "item_weight": current.get("item_weight"),
        "item_weight_unit": current.get("item_weight_unit"),
        "item_length": current.get("item_length"),
        "item_width": current.get("item_width"),
        "item_height": current.get("item_height"),
        "item_dimensions_unit": current.get("item_dimensions_unit"),
        "tags": tags,
        "materials": materials,
        "images": images,
        "weight": weight_display,
        "dimensions": dimensions_display,
    }

    # Informational (non-required) shipping fields shown in the comparison table
    info_fields = [
        {"key": "weight", "label": "Weight"},
        {"key": "dimensions", "label": "Dimensions (L×W×H)"},
        {"key": "tags", "label": "Tags"},
    ]

    required_fields = []
    missing_current: list[str] = []
    missing_optimized: list[str] = []
    for key, label in ETSY_REQUIRED_CREATE_FIELDS:
        cur_val = current.get(key)
        opt_val = optimized.get(key)
        cur_missing = _is_missing_etsy_required(key, cur_val, listing_type=listing_type_current)
        opt_missing = _is_missing_etsy_required(key, opt_val, listing_type=optimized_type)
        if cur_missing:
            missing_current.append(key)
        if opt_missing:
            missing_optimized.append(key)

        required_fields.append(
            {
                "key": key,
                "label": label,
                "current": cur_val,
                "optimized": opt_val,
                "current_missing": cur_missing,
                "optimized_missing": opt_missing,
                "informational": False,
            }
        )

    for info in info_fields:
        key = info["key"]
        val = current.get(key)
        required_fields.append(
            {
                "key": key,
                "label": info["label"],
                "current": val,
                "optimized": val,
                "current_missing": val is None,
                "optimized_missing": val is None,
                "informational": True,
            }
        )

    return {
        "sku": doc.get("_id") or doc.get("sku"),
        "title": doc.get("title"),
        "category": doc.get("category"),
        "etsy_linked": bool((etsy.get("listing_id") or "").strip()) if isinstance(etsy.get("listing_id"), str) else bool(etsy.get("listing_id")),
        "etsy_listing_id": etsy.get("listing_id"),
        "last_normalized_at": doc.get("last_normalized_at"),
        "image": images[0] if images else None,
        "current": current,
        "etsy_optimized": optimized,
        "required_fields": required_fields,
        "required_missing_current": missing_current,
        "required_missing_optimized": missing_optimized,
        "required_ready_current": len(missing_current) == 0,
        "required_ready_optimized": len(missing_optimized) == 0,
        "package": {
            "weight_display": weight_display,
            "dimensions_display": dimensions_display,
            "weight": package.get("weight"),
            "dimensions": package.get("dimensions"),
        },
    }


def _has_required_shipping_measurements(package: dict | None) -> bool:
    """Check if physical item has weight and dimensions for Etsy."""
    if not isinstance(package, dict):
        return False
    
    weight = package.get("weight") or {}
    if not isinstance(weight, dict):
        return False
    
    major = weight.get("major") or {}
    major_value = _to_float(major.get("value"))
    if major_value is None or major_value <= 0:
        return False
    
    dimensions = package.get("dimensions") or {}
    if not isinstance(dimensions, dict):
        return False
    
    has_all_dims = all(
        isinstance(dimensions.get(key), dict) and _to_float(dimensions[key].get("value")) is not None
        for key in ["length", "width", "height"]
    )
    return has_all_dims


async def _generate_etsy_optimizations_for_bulk(detail: dict) -> dict[str, Any]:
    """Generate all Etsy optimizations (SEO, tags, taxonomy) for bulk publishing."""
    optimizations: dict[str, Any] = {
        "generated_at": _now_utc(),
    }
    
    # Generate SEO optimization (title + when_made)
    seo_opt = await _generate_etsy_seo_optimization(detail)
    optimizations["seo_title"] = seo_opt.get("seo_title")
    optimizations["when_made"] = seo_opt.get("when_made")
    optimizations["seo_ai_used"] = seo_opt.get("ai_used")
    optimizations["seo_reason"] = seo_opt.get("reason")
    
    # Update detail with optimized values for downstream functions
    detail["etsy_optimized"]["title"] = seo_opt.get("seo_title") or detail["etsy_optimized"].get("title")
    detail["etsy_optimized"]["when_made"] = seo_opt.get("when_made") or detail["etsy_optimized"].get("when_made")
    
    # Generate tag suggestions
    tag_opt = await _generate_etsy_tag_suggestions(detail)
    optimizations["tags"] = tag_opt.get("tags") or []
    optimizations["tags_ai_used"] = tag_opt.get("ai_used")
    optimizations["tags_reason"] = tag_opt.get("reason")
    
    # Update detail for taxonomy function
    detail["etsy_optimized"]["tags"] = optimizations["tags"]
    
    # Suggest taxonomy
    taxonomy_opt = await _suggest_etsy_taxonomy(detail, limit=5)
    ai_choice = taxonomy_opt.get("ai_choice") or {}
    optimizations["taxonomy_id"] = ai_choice.get("taxonomy_id")
    optimizations["taxonomy_confidence"] = ai_choice.get("confidence") or 0.0
    optimizations["taxonomy_ai_used"] = taxonomy_opt.get("ai_used")
    optimizations["taxonomy_reason"] = ai_choice.get("reason")
    optimizations["taxonomy_best_match"] = taxonomy_opt.get("best_match")
    optimizations["taxonomy_suggestions"] = taxonomy_opt.get("suggestions") or []
    
    return optimizations


async def _validate_bulk_item(
    doc: dict,
    *,
    min_taxonomy_confidence: float = ETSY_BULK_MIN_TAXONOMY_CONFIDENCE,
) -> dict[str, Any]:
    """Validate a single item for bulk publishing. Generates optimizations first, then validates."""
    sku = doc.get("_id") or doc.get("sku")
    
    # Early exit checks (don't need optimization for these)
    existing_listing_id = ((doc.get("channels") or {}).get("etsy") or {}).get("listing_id")
    if existing_listing_id:
        return {
            "sku": sku,
            "status": "skipped",
            "reason": "already_linked_to_etsy",
            "detail": f"Existing listing_id: {existing_listing_id}",
            "optimizations": None,
        }
    
    qty = _to_int(doc.get("quantity"))
    if qty is None or qty <= 0:
        return {
            "sku": sku,
            "status": "skipped",
            "reason": "insufficient_quantity",
            "detail": f"Quantity: {qty}",
            "optimizations": None,
        }
    
    # Build comparison to get base view
    detail = _build_etsy_publish_comparison(doc)
    
    # GENERATE OPTIMIZATIONS FIRST
    try:
        optimizations = await _generate_etsy_optimizations_for_bulk(detail)
    except Exception as opt_exc:
        return {
            "sku": sku,
            "status": "skipped",
            "reason": "optimization_generation_failed",
            "detail": f"Failed to generate optimizations: {str(opt_exc)}",
            "optimizations": None,
        }
    
    # Now validate using optimized data
    optimized = detail.get("etsy_optimized") or {}
    
    # Check required fields (using optimized payload)
    test_payload = _build_etsy_create_listing_payload(detail)
    missing_fields = _get_missing_etsy_required_fields(test_payload)
    if missing_fields:
        return {
            "sku": sku,
            "status": "skipped",
            "reason": "missing_required_fields",
            "detail": f"Missing: {', '.join(missing_fields)}",
            "optimizations": optimizations,
        }
    
    # For physical listings, check shipping measurements
    listing_type = _coerce_etsy_enum(optimized.get("type"), ETSY_ALLOWED_LISTING_TYPES) or settings.ETSY_DEFAULT_LISTING_TYPE
    if listing_type == "physical":
        package = doc.get("package")
        if not _has_required_shipping_measurements(package):
            return {
                "sku": sku,
                "status": "skipped",
                "reason": "missing_shipping_measurements",
                "detail": "Physical listing requires weight and all three dimensions (length, width, height)",
                "optimizations": optimizations,
            }
    
    # Check taxonomy confidence (generated in optimizations)
    taxonomy_id = optimizations.get("taxonomy_id")
    taxonomy_confidence = optimizations.get("taxonomy_confidence") or 0.0
    
    if not taxonomy_id or taxonomy_confidence < min_taxonomy_confidence:
        reason = "low_confidence_taxonomy" if not taxonomy_id else "low_confidence_taxonomy_threshold"
        return {
            "sku": sku,
            "status": "skipped",
            "reason": reason,
            "detail": f"AI taxonomy confidence: {taxonomy_confidence} (threshold: {min_taxonomy_confidence})",
            "optimizations": optimizations,
        }
    
    # All checks passed - item is ready
    return {
        "sku": sku,
        "status": "ready",
        "reason": None,
        "detail": None,
        "optimizations": optimizations,
    }


async def _generate_bulk_report(
    validation_results: list[dict],
    creation_results: list[dict] | None = None,
) -> dict[str, Any]:
    """Generate a comprehensive bulk report with optimization details."""
    report: dict[str, Any] = {
        "generated_at": _now_utc(),
        "validation": {
            "total_checked": len(validation_results),
            "ready_to_create": sum(1 for r in validation_results if r.get("status") == "ready"),
            "skipped": sum(1 for r in validation_results if r.get("status") == "skipped"),
        },
        "validation_items": validation_results,
    }
    
    if creation_results:
        report["creation"] = {
            "total_attempted": len(creation_results),
            "created": sum(1 for r in creation_results if r.get("status") == "created"),
            "failed": sum(1 for r in creation_results if r.get("status") == "failed"),
        }
        report["creation_items"] = creation_results
    
    # Generate list of items needing data fixes
    needs_weight_dims = [
        {
            "sku": r.get("sku"),
            "reason": r.get("reason"),
            "detail": r.get("detail"),
        }
        for r in validation_results
        if r.get("status") == "skipped" and r.get("reason") == "missing_shipping_measurements"
    ]
    if needs_weight_dims:
        report["items_needing_data_fixes"] = {
            "category": "missing_shipping_measurements",
            "count": len(needs_weight_dims),
            "items": needs_weight_dims,
        }
    
    # Optimization summary (only for ready items)
    ready_items = [r for r in validation_results if r.get("status") == "ready"]
    if ready_items:
        report["optimization_summary"] = {
            "total_optimized": len(ready_items),
            "ai_seo_used": sum(1 for r in ready_items if (r.get("optimizations") or {}).get("seo_ai_used")),
            "ai_tags_used": sum(1 for r in ready_items if (r.get("optimizations") or {}).get("tags_ai_used")),
            "ai_taxonomy_used": sum(1 for r in ready_items if (r.get("optimizations") or {}).get("taxonomy_ai_used")),
        }
    
    return report


async def _fetch_shopify_snapshot(shopify_id: object) -> dict | None:
    sid = _to_int(shopify_id)
    if sid is None:
        return None

    client = None
    if settings.SHOPIFY_API_KEY_PROD and settings.SHOPIFY_PASSWORD_PROD and settings.SHOPIFY_STORE_URL_PROD:
        client = ShopifyClient(
            api_key=settings.SHOPIFY_API_KEY_PROD,
            password=settings.SHOPIFY_PASSWORD_PROD,
            store_url=settings.SHOPIFY_STORE_URL_PROD,
        )

    if client is None:
        return {
            "product_id": sid,
            "error": "shopify_prod_credentials_not_configured",
        }

    product = None
    source_store = None
    last_error = None
    try:
        resp = await client.get(f"products/{sid}.json")
        candidate = (resp or {}).get("product")
        if isinstance(candidate, dict):
            product = candidate
            source_store = "prod"
    except Exception as exc:
        last_error = str(exc)

    if not isinstance(product, dict):
        return {
            "product_id": sid,
            "error": last_error or "shopify_product_not_found",
        }

    variants = product.get("variants") or []
    first_variant = variants[0] if variants else {}
    images = product.get("images") or []
    image_src = None
    if images and isinstance(images[0], dict):
        image_src = images[0].get("src")

    return {
        "product_id": sid,
        "store": source_store,
        "variant_id": first_variant.get("id") if isinstance(first_variant, dict) else None,
        "title": product.get("title"),
        "status": product.get("status"),
        "handle": product.get("handle"),
        "price": _to_float(first_variant.get("price") if isinstance(first_variant, dict) else None),
        "quantity": _to_int(first_variant.get("inventory_quantity") if isinstance(first_variant, dict) else None),
        "image": image_src,
    }


async def _apply_match_row(match_row: dict, *, matched_by: str) -> tuple[bool, str]:
    listing_id = match_row.get("etsy_listing_id")
    sku = match_row.get("normalized_sku")
    if not listing_id or not sku:
        return False, "missing_listing_or_sku"

    etsy_doc = await db.etsy_listings_investigation.find_one(
        {"listing_id": listing_id},
        {
            "listing_id": 1,
            "listing_state": 1,
            "title": 1,
            "url": 1,
            "shop_id": 1,
            "price": 1,
            "quantity": 1,
        },
    )
    if not etsy_doc:
        return False, "etsy_listing_not_found"

    channels_etsy = {
        "listing_id": etsy_doc.get("listing_id"),
        "listing_state": etsy_doc.get("listing_state"),
        "title": etsy_doc.get("title"),
        "url": etsy_doc.get("url"),
        "shop_id": etsy_doc.get("shop_id"),
        "price": etsy_doc.get("price"),
        "quantity": etsy_doc.get("quantity"),
        "match": {
            "status": "approved",
            "matched_by": matched_by,
            "score": match_row.get("score"),
            "bucket": match_row.get("bucket"),
            "matched_at": _now_utc(),
        },
    }

    result = await db.product_normalized.update_one(
        {"_id": sku},
        {
            "$set": {
                "channels.etsy": channels_etsy,
            }
        },
    )
    if result.matched_count == 0:
        return False, "normalized_sku_not_found"

    await db[REVIEW_COLLECTION].update_one(
        {"etsy_listing_id": listing_id, "normalized_sku": sku},
        {
            "$set": {
                "status": "approved",
                "bucket": match_row.get("bucket"),
                "score": match_row.get("score"),
                "matched_by": matched_by,
                "updated_at": _now_utc(),
            }
        },
        upsert=True,
    )

    return True, "applied"


@router.get("/items")
async def report_items(
    q: str | None = Query(default=None, description="Search across SKU / title / item id"),
    size: str | None = Query(default=None, description="Size attribute (string contains match)"),
    posted_after: datetime | None = Query(default=None, description="eBay listing start time >= this (ISO 8601)"),
    posted_before: datetime | None = Query(default=None, description="eBay listing start time <= this (ISO 8601)"),
    min_weight: float | None = Query(default=None, description="Min package weight major value"),
    max_weight: float | None = Query(default=None, description="Max package weight major value"),
    min_length: float | None = Query(default=None, description="Min package length"),
    max_length: float | None = Query(default=None, description="Max package length"),
    min_width: float | None = Query(default=None, description="Min package width"),
    max_width: float | None = Query(default=None, description="Max package width"),
    min_height: float | None = Query(default=None, description="Min package height"),
    max_height: float | None = Query(default=None, description="Max package height"),
    available_only: bool = Query(default=False, description="Only items with quantity > 0"),
    soldout_only: bool = Query(default=False, description="Only items with quantity <= 0"),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
):
    posted_after = _ensure_utc(posted_after)
    posted_before = _ensure_utc(posted_before)

    and_conditions: list[dict] = []

    # Posting date filter (stored on product_raw)
    if posted_after or posted_before:
        date_q: dict[str, object] = {}
        if posted_after:
            date_q["$gte"] = posted_after
        if posted_before:
            date_q["$lte"] = posted_before
        and_conditions.append({"ebay_posted_at": date_q})

    # Search query
    if q:
        rx = _rx(q)
        and_conditions.append(
            {
                "$or": [
                    {"_id": rx},
                    {"raw.ItemID": rx},
                    {"raw.SKU": rx},
                    {"raw.Title": rx},
                    {"normalized.title": rx},
                    {"normalized.category": rx},
                ]
            }
        )

    # Size filter (best-effort)
    if size:
        rx = _rx(size)
        and_conditions.append(
            {
                "$or": [
                    {"raw.ItemSpecifics.Size": rx},
                    {"normalized.attributes.Size": rx},
                    {"normalized.metafields.art.size": rx},
                ]
            }
        )

    # Weight filter (normalized.package.weight.major.value)
    if min_weight is not None or max_weight is not None:
        wq: dict[str, object] = {}
        if min_weight is not None:
            wq["$gte"] = float(min_weight)
        if max_weight is not None:
            wq["$lte"] = float(max_weight)
        and_conditions.append({"normalized.package.weight.major.value": wq})

    # Dimensions filters
    def _dim_cond(field: str, min_v: float | None, max_v: float | None) -> None:
        if min_v is None and max_v is None:
            return
        dq: dict[str, object] = {}
        if min_v is not None:
            dq["$gte"] = float(min_v)
        if max_v is not None:
            dq["$lte"] = float(max_v)
        and_conditions.append({field: dq})

    _dim_cond("normalized.package.dimensions.length.value", min_length, max_length)
    _dim_cond("normalized.package.dimensions.width.value", min_width, max_width)
    _dim_cond("normalized.package.dimensions.height.value", min_height, max_height)

    # Availability filter (prefer normalized.quantity, fall back to raw.QuantityAvailable)
    # If both checkboxes are set (or neither), treat as no filter.
    if available_only and not soldout_only:
        and_conditions.append(
            {
                "$expr": {
                    "$gt": [
                        {"$ifNull": ["$normalized.quantity", "$raw.QuantityAvailable"]},
                        0,
                    ]
                }
            }
        )
    elif soldout_only and not available_only:
        and_conditions.append(
            {
                "$expr": {
                    "$lte": [
                        {"$ifNull": ["$normalized.quantity", "$raw.QuantityAvailable"]},
                        0,
                    ]
                }
            }
        )

    match_stage: dict | None = None
    if and_conditions:
        match_stage = {"$match": {"$and": and_conditions}}

    lookup = {
        "$lookup": {
            "from": "product_normalized",
            "localField": "_id",
            "foreignField": "_id",
            "as": "normalized",
        }
    }
    unwind = {"$unwind": {"path": "$normalized", "preserveNullAndEmptyArrays": True}}

    base_pipeline: list[dict] = [lookup, unwind]
    if match_stage:
        base_pipeline.append(match_stage)

    # Count pipeline (same filters)
    count_pipeline = base_pipeline + [{"$count": "total"}]
    count_docs = await db.product_raw.aggregate(count_pipeline).to_list(length=1)
    total = int(count_docs[0]["total"]) if count_docs else 0

    # Result pipeline
    result_pipeline = (
        base_pipeline
        + [
            {"$sort": {"ebay_posted_at": -1, "_id": 1}},
            {"$skip": int(skip)},
            {"$limit": int(limit)},
            {
                "$project": {
                    "_id": 0,
                    "sku": "$_id",
                    "ebay_posted_at": 1,
                    "image_url": {
                        "$ifNull": [
                            {"$arrayElemAt": ["$normalized.images", 0]},
                            {"$arrayElemAt": ["$raw.Images", 0]},
                        ]
                    },
                    "image_urls": {"$ifNull": ["$normalized.images", "$raw.Images"]},
                    "raw": "$raw",
                    "normalized": "$normalized",
                }
            },
        ]
    )

    items = await db.product_raw.aggregate(result_pipeline).to_list(length=limit)

    return {
        "items": items,
        "total": total,
        "skip": skip,
        "limit": limit,
        "has_more": (skip + len(items)) < total,
    }


@router.get("/etsy-match/summary")
async def etsy_match_summary():
    """Get current Etsy match summary from database (live counts)."""
    # Count by bucket/status from review collection
    pipeline = [
        {
            '$group': {
                '_id': {'bucket': '$bucket', 'status': '$status'},
                'count': {'$sum': 1}
            }
        }
    ]
    review_breakdown = await db.etsy_match_review.aggregate(pipeline).to_list(None)
    
    # Count linked products
    linked = await db.product_normalized.count_documents({'channels.etsy.listing_id': {'$exists': True}})
    
    # Total unmatched
    total_etsy = await db.etsy_listings_investigation.count_documents({})
    unmatched = total_etsy - linked
    
    # Build summary object
    summary = {
        'etsy_total': total_etsy,
        'linked': linked,
        'unmatched': unmatched,
    }
    
    # Extract approved counts by bucket
    for row in review_breakdown:
        bucket = row['_id']['bucket']
        status = row['_id']['status']
        count = row['count']
        
        if status == 'approved':
            if bucket == 'exact_ci':
                summary['exact'] = count
            elif bucket == 'normalized_exact':
                summary['normalized_exact'] = count
            elif bucket == 'high_confidence':
                summary['high'] = count
            elif bucket == 'medium_confidence':
                summary['medium'] = count
            elif bucket == 'low_confidence':
                summary['low'] = count
    
    # Count pending by bucket for UI
    pending_counts = {}
    for row in review_breakdown:
        bucket = row['_id']['bucket']
        status = row['_id']['status']
        count = row['count']
        
        if status == 'pending':
            pending_counts[bucket] = count
    
    summary['pending'] = pending_counts
    
    return {
        "summary": summary,
        "generated_at": _now_utc(),
    }


@router.post("/etsy-match/apply-exact")
async def etsy_match_apply_exact():
    report = _load_match_report()
    rows = (report.get("exact_matches") or []) + (report.get("normalized_exact_matches") or [])

    applied = 0
    skipped = 0
    errors: list[dict] = []

    for row in rows:
        ok, reason = await _apply_match_row(row, matched_by="auto_exact")
        if ok:
            applied += 1
        else:
            skipped += 1
            errors.append(
                {
                    "etsy_listing_id": row.get("etsy_listing_id"),
                    "normalized_sku": row.get("normalized_sku"),
                    "reason": reason,
                }
            )

    return {
        "matched_candidates": len(rows),
        "applied": applied,
        "skipped": skipped,
        "errors": errors[:50],
    }


@router.get("/etsy-match/review")
async def etsy_match_review(
    bucket: str = Query(default="all", description="all|high|medium|low|unmatched"),
    status: str = Query(default="pending", description="pending|approved|denied|all"),
    limit: int = Query(default=200, ge=1, le=1000),
):
    report = _load_match_report()

    bucket_map = {
        "high": "high_confidence_matches",
        "medium": "medium_confidence_matches",
        "low": "low_confidence_matches",
        "unmatched": "unmatched",
    }

    if bucket == "all":
        source_rows = (
            (report.get("high_confidence_matches") or [])
            + (report.get("medium_confidence_matches") or [])
            + (report.get("low_confidence_matches") or [])
            + (report.get("unmatched") or [])
        )
    else:
        key = bucket_map.get(bucket)
        if not key:
            raise HTTPException(status_code=400, detail="Invalid bucket")
        source_rows = report.get(key) or []

    out: list[dict] = []
    for row in source_rows:
        listing_id = row.get("etsy_listing_id")
        sku = row.get("normalized_sku")

        review_doc = await db[REVIEW_COLLECTION].find_one(
            {"etsy_listing_id": listing_id, "normalized_sku": sku},
            {"_id": 0, "status": 1, "reason": 1, "updated_at": 1},
        )
        row_status = (review_doc or {}).get("status", "pending")

        if status != "all" and row_status != status:
            continue

        etsy_doc = None
        if listing_id is not None:
            etsy_doc = await db.etsy_listings_investigation.find_one(
                {"listing_id": listing_id},
                {"_id": 0, "price": 1, "quantity": 1, "listing_state": 1},
            )

        normalized_doc = None
        if sku:
            normalized_doc = await db.product_normalized.find_one(
                {"_id": sku},
                {
                    "_id": 1,
                    "price": 1,
                    "shipping": 1,
                    "shopify_id": 1,
                    "channels.shopify.shopify_id": 1,
                },
            )

        shopify_id = None
        if normalized_doc:
            shopify_id = (
                (normalized_doc.get("channels") or {}).get("shopify", {}).get("shopify_id")
                or normalized_doc.get("shopify_id")
            )

        out.append(
            {
                **row,
                "review_status": row_status,
                "review_reason": (review_doc or {}).get("reason"),
                "review_updated_at": (review_doc or {}).get("updated_at"),
                "etsy_link": f"https://www.etsy.com/listing/{listing_id}" if listing_id else None,
                "shopify_id": shopify_id,
                "shopify_links": _build_shopify_links(shopify_id),
                "normalized_price": (normalized_doc or {}).get("price"),
                "normalized_shipping": (normalized_doc or {}).get("shipping"),
                "etsy_price": (etsy_doc or {}).get("price"),
                "etsy_state_live": (etsy_doc or {}).get("listing_state"),
            }
        )

        if len(out) >= limit:
            break

    return {
        "bucket": bucket,
        "status": status,
        "count": len(out),
        "items": out,
    }


@router.post("/etsy-match/review/approve")
async def etsy_match_approve(payload: dict):
    listing_id = payload.get("etsy_listing_id")
    sku = payload.get("normalized_sku")
    if not listing_id or not sku:
        raise HTTPException(status_code=400, detail="etsy_listing_id and normalized_sku are required")

    row = {
        "etsy_listing_id": listing_id,
        "normalized_sku": sku,
        "score": payload.get("score"),
        "bucket": payload.get("bucket"),
    }
    ok, reason = await _apply_match_row(row, matched_by="manual_review")
    if not ok:
        raise HTTPException(status_code=400, detail=reason)
    return {"ok": True}


@router.post("/etsy-match/review/deny")
async def etsy_match_deny(payload: dict):
    listing_id = payload.get("etsy_listing_id")
    sku = payload.get("normalized_sku")
    reason = payload.get("reason")
    if not listing_id:
        raise HTTPException(status_code=400, detail="etsy_listing_id is required")

    await db[REVIEW_COLLECTION].update_one(
        {"etsy_listing_id": listing_id, "normalized_sku": sku},
        {
            "$set": {
                "status": "denied",
                "reason": reason,
                "updated_at": _now_utc(),
            }
        },
        upsert=True,
    )
    return {"ok": True}


@router.get("/etsy-match-excel/summary")
async def etsy_match_excel_summary():
    workbook_path = _find_latest_excel_review_path()
    rows, _ = _load_excel_rows(workbook_path)
    with_shopify = sum(1 for row in rows if row.get("matched_shopify_link"))
    with_sku = sum(1 for row in rows if row.get("matched_mongo_sku"))
    return {
        "workbook": str(workbook_path),
        "count": len(rows),
        "with_shopify_link": with_shopify,
        "with_matched_sku": with_sku,
        "generated_at": _now_utc(),
    }


@router.get("/etsy-match-excel/review")
async def etsy_match_excel_review(
    limit: int = Query(default=500, ge=1, le=2000),
):
    workbook_path = _find_latest_excel_review_path()
    rows, _ = _load_excel_rows(workbook_path)
    selected = rows[:limit]

    listing_ids: list[int] = []
    skus: list[str] = []
    for row in selected:
        try:
            listing_ids.append(int(row.get("listing_id")))
        except Exception:
            pass
        sku = row.get("matched_mongo_sku")
        if sku:
            skus.append(str(sku))

    etsy_docs_by_listing: dict[int, dict] = {}
    if listing_ids:
        cursor = db.etsy_listings_investigation.find(
            {"listing_id": {"$in": listing_ids}},
            {"_id": 0, "listing_id": 1, "title": 1, "url": 1, "raw.images": 1},
        )
        async for doc in cursor:
            try:
                etsy_docs_by_listing[int(doc.get("listing_id"))] = doc
            except Exception:
                continue

    norm_docs_by_sku: dict[str, dict] = {}
    if skus:
        cursor = db.product_normalized.find(
            {"_id": {"$in": skus}},
            {
                "_id": 1,
                "title": 1,
                "images": 1,
                "quantity": 1,
                "shopify_id": 1,
                "channels.shopify.shopify_id": 1,
            },
        )
        async for doc in cursor:
            norm_docs_by_sku[str(doc.get("_id"))] = doc

    etsy_images_by_listing: dict[int, str | None] = {}
    listings_missing_image: list[int] = []

    for listing_id in listing_ids:
        etsy_doc = etsy_docs_by_listing.get(listing_id)
        image = _extract_etsy_image_from_doc(etsy_doc)
        etsy_images_by_listing[listing_id] = image
        if not image:
            listings_missing_image.append(listing_id)

    if listings_missing_image:
        headers = None
        try:
            headers = await _resolve_etsy_auth_headers_for_review()
        except HTTPException:
            headers = None

        if headers:
            semaphore = asyncio.Semaphore(6)
            async with httpx.AsyncClient(timeout=20.0) as client:
                async def _fetch_one(lid: int) -> None:
                    async with semaphore:
                        image_url = await _fetch_etsy_main_image_from_api(client, headers, lid)
                        if image_url:
                            etsy_images_by_listing[lid] = image_url

                await asyncio.gather(*(_fetch_one(lid) for lid in listings_missing_image))

    out: list[dict] = []
    for row in selected:
        listing_id_raw = row.get("listing_id")
        try:
            listing_id = int(listing_id_raw)
        except Exception:
            listing_id = None

        sku = str(row.get("matched_mongo_sku") or "").strip() or None
        etsy_doc = etsy_docs_by_listing.get(listing_id) if listing_id is not None else None
        norm_doc = norm_docs_by_sku.get(sku) if sku else None

        shopify_id = None
        if norm_doc:
            shopify_id = (
                (norm_doc.get("channels") or {}).get("shopify", {}).get("shopify_id")
                or norm_doc.get("shopify_id")
            )

        shopify_image = None
        if norm_doc:
            images = norm_doc.get("images") or []
            if images:
                shopify_image = images[0]

        etsy_image = etsy_images_by_listing.get(listing_id) if listing_id is not None else None

        out.append(
            {
                "listing_id": listing_id,
                "etsy_title": row.get("live_title") or (etsy_doc or {}).get("title"),
                "etsy_link": row.get("live_url") or (etsy_doc or {}).get("url"),
                "etsy_image": etsy_image,
                "matched_mongo_title": row.get("matched_mongo_title") or (norm_doc or {}).get("title"),
                "matched_mongo_sku": sku,
                "confidence_rate": row.get("confidence_rate"),
                "matched_shopify_link": row.get("matched_shopify_link"),
                "shopify_image": shopify_image,
                "shopify_id": shopify_id,
                "shopify_links": _build_shopify_links(shopify_id),
                "mongo_quantity": (norm_doc or {}).get("quantity"),
            }
        )

    return {
        "workbook": str(workbook_path),
        "count": len(out),
        "items": out,
    }


@router.post("/etsy-match-excel/review/approve")
async def etsy_match_excel_approve(payload: dict):
    return await _approve_excel_match(payload)


async def _approve_excel_match(
    payload: dict,
    *,
    headers: dict[str, str] | None = None,
    workbook_path: Path | None = None,
) -> dict:
    listing_id = payload.get("listing_id")
    sku = str(payload.get("matched_mongo_sku") or "").strip()
    if not listing_id or not sku:
        raise HTTPException(status_code=400, detail="listing_id and matched_mongo_sku are required")

    try:
        listing_id_int = int(listing_id)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid_listing_id") from exc

    norm_doc = await db.product_normalized.find_one({"_id": sku}, {"_id": 1, "quantity": 1})
    if not norm_doc:
        raise HTTPException(status_code=404, detail="matched_mongo_sku_not_found")

    etsy_headers = headers or await _resolve_etsy_auth_headers_for_review()
    target_quantity = int(norm_doc.get("quantity") or 0)
    await _update_etsy_listing_sku_for_review(
        listing_id=listing_id_int,
        new_sku=sku,
        target_quantity=target_quantity,
        headers=etsy_headers,
    )

    ok, reason = await _apply_match_row(
        {
            "etsy_listing_id": listing_id_int,
            "normalized_sku": sku,
            "score": payload.get("confidence_rate"),
            "bucket": "excel_review",
        },
        matched_by="excel_manual_review",
    )
    if not ok:
        raise HTTPException(status_code=400, detail=reason)

    workbook = workbook_path or _find_latest_excel_review_path()
    removed = _remove_listing_from_excel(workbook, listing_id_int)

    return {
        "ok": True,
        "listing_id": listing_id_int,
        "matched_mongo_sku": sku,
        "removed_from_excel": removed,
        "workbook": str(workbook),
    }


@router.put("/etsy-match-excel/review/approve-batch")
async def etsy_match_excel_approve_batch(payload: dict):
    approvals = payload.get("approvals") if isinstance(payload, dict) else None
    if not isinstance(approvals, list) or not approvals:
        raise HTTPException(status_code=400, detail="approvals list is required")

    workbook_path = _find_latest_excel_review_path()
    headers = await _resolve_etsy_auth_headers_for_review()

    successes: list[dict] = []
    failures: list[dict] = []
    for item in approvals:
        try:
            result = await _approve_excel_match(item, headers=headers, workbook_path=workbook_path)
            successes.append(
                {
                    "listing_id": result.get("listing_id"),
                    "matched_mongo_sku": result.get("matched_mongo_sku"),
                    "removed_from_excel": result.get("removed_from_excel"),
                }
            )
        except HTTPException as exc:
            failures.append(
                {
                    "listing_id": item.get("listing_id") if isinstance(item, dict) else None,
                    "matched_mongo_sku": (item or {}).get("matched_mongo_sku") if isinstance(item, dict) else None,
                    "detail": exc.detail,
                    "status_code": exc.status_code,
                }
            )
        except Exception as exc:
            failures.append(
                {
                    "listing_id": item.get("listing_id") if isinstance(item, dict) else None,
                    "matched_mongo_sku": (item or {}).get("matched_mongo_sku") if isinstance(item, dict) else None,
                    "detail": str(exc),
                    "status_code": 500,
                }
            )

    return {
        "ok": len(failures) == 0,
        "workbook": str(workbook_path),
        "requested": len(approvals),
        "applied": len(successes),
        "failed": len(failures),
        "successes": successes,
        "failures": failures,
    }


@router.post("/etsy-match-excel/review/deny")
async def etsy_match_excel_deny(payload: dict):
    listing_id = payload.get("listing_id")
    sku = str(payload.get("matched_mongo_sku") or "").strip() or None
    reason = payload.get("reason")
    if not listing_id:
        raise HTTPException(status_code=400, detail="listing_id is required")

    await db[REVIEW_COLLECTION].update_one(
        {"etsy_listing_id": listing_id, "normalized_sku": sku},
        {
            "$set": {
                "status": "denied",
                "reason": reason,
                "bucket": "excel_review",
                "updated_at": _now_utc(),
            }
        },
        upsert=True,
    )
    return {"ok": True}


@router.get("/etsy-publish/queue")
async def etsy_publish_queue(
    q: str | None = Query(default=None, description="Search SKU/title/category"),
    limit: int = Query(default=100, ge=1, le=500),
):
    query: dict = {
        "quantity": {"$gte": 1},
        "$or": [
            {"channels.etsy.listing_id": {"$exists": False}},
            {"channels.etsy.listing_id": None},
            {"channels.etsy.listing_id": ""},
        ]
    }
    if q:
        rx = _rx(q)
        query["$and"] = [
            {
                "$or": [
                    {"_id": rx},
                    {"sku": rx},
                    {"title": rx},
                    {"category": rx},
                ]
            }
        ]

    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }

    total = await db.product_normalized.count_documents(query)
    docs = await db.product_normalized.find(query, projection).sort("last_normalized_at", -1).to_list(length=limit)

    items: list[dict] = []
    for doc in docs:
        detail = _build_etsy_publish_comparison(doc)
        items.append(
            {
                "sku": detail.get("sku"),
                "title": detail.get("title"),
                "category": doc.get("category"),
                "image": detail.get("image"),
                "required_missing_current": detail.get("required_missing_current"),
                "required_missing_optimized": detail.get("required_missing_optimized"),
                "required_ready_current": detail.get("required_ready_current"),
                "required_ready_optimized": detail.get("required_ready_optimized"),
                "last_normalized_at": detail.get("last_normalized_at"),
            }
        )

    return {
        "count": len(items),
        "total": int(total),
        "items": items,
    }


@router.get("/etsy-publish/{sku}")
async def etsy_publish_detail(sku: str):
    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }
    doc = await db.product_normalized.find_one({"$or": [{"_id": sku}, {"sku": sku}]}, projection)
    if not doc:
        raise HTTPException(status_code=404, detail="SKU not found")

    detail = _build_etsy_publish_comparison(doc)

    seo_optimization = await _generate_etsy_seo_optimization(detail)
    detail["etsy_optimized"]["title"] = seo_optimization.get("seo_title") or detail["etsy_optimized"].get("title")
    detail["etsy_optimized"]["when_made"] = seo_optimization.get("when_made") or detail["etsy_optimized"].get("when_made")
    detail["seo_optimization"] = seo_optimization

    for row in detail.get("required_fields") or []:
        if row.get("key") == "title":
            row["optimized"] = detail["etsy_optimized"].get("title")
        elif row.get("key") == "when_made":
            row["optimized"] = detail["etsy_optimized"].get("when_made")

    tag_suggestions = await _generate_etsy_tag_suggestions(detail)
    detail["current"]["tags"] = tag_suggestions.get("source_tags") or detail["current"].get("tags") or []
    detail["etsy_optimized"]["tags"] = tag_suggestions.get("tags") or detail["etsy_optimized"].get("tags") or []
    detail["tag_suggestions"] = tag_suggestions
    for row in detail.get("required_fields") or []:
        if row.get("key") == "tags":
            row["current"] = detail["current"].get("tags") or []
            row["optimized"] = detail["etsy_optimized"].get("tags") or []
    detail["required_field_help"] = {
        "type": sorted(ETSY_ALLOWED_LISTING_TYPES),
        "who_made": sorted(ETSY_ALLOWED_WHO_MADE),
        "when_made": sorted(ETSY_ALLOWED_WHEN_MADE),
    }
    return detail


@router.get("/etsy-publish/{sku}/suggest-taxonomy")
async def etsy_publish_suggest_taxonomy(sku: str, limit: int = Query(default=ETSY_TAXONOMY_RESULT_LIMIT, ge=1, le=20)):
    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }
    doc = await db.product_normalized.find_one({"$or": [{"_id": sku}, {"sku": sku}]}, projection)
    if not doc:
        raise HTTPException(status_code=404, detail="SKU not found")

    detail = _build_etsy_publish_comparison(doc)
    suggestion_data = await _suggest_etsy_taxonomy(detail, limit=limit)
    return {
        "ok": True,
        "sku": detail.get("sku"),
        "query": suggestion_data.get("query", {}).get("query_text"),
        "ai_used": suggestion_data.get("ai_used", False),
        "ai_choice": suggestion_data.get("ai_choice"),
        "best_match": suggestion_data.get("best_match"),
        "suggestions": suggestion_data.get("suggestions") or [],
    }


@router.post("/etsy-publish/{sku}/create-draft")
async def etsy_publish_create_draft(sku: str, body: dict | None = Body(default=None)):
    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }
    doc = await db.product_normalized.find_one({"$or": [{"_id": sku}, {"sku": sku}]}, projection)
    if not doc:
        raise HTTPException(status_code=404, detail="SKU not found")

    existing_listing_id = ((doc.get("channels") or {}).get("etsy") or {}).get("listing_id")
    if existing_listing_id:
        sku_value = str(doc.get("_id") or doc.get("sku") or sku)
        await _record_etsy_draft_attempt(
            sku=sku_value,
            shop_id=str(((doc.get("channels") or {}).get("etsy") or {}).get("shop_id") or "") or None,
            request_payload=None,
            request_form_data=None,
            response_status_code=None,
            response_payload={
                "ok": False,
                "error": "already_linked_to_etsy",
                "existing_listing_id": existing_listing_id,
            },
            ok=False,
            linked_listing_id=_to_int(existing_listing_id),
            error_code="already_linked_to_etsy",
        )
        return {
            "ok": False,
            "error": "already_linked_to_etsy",
            "sku": doc.get("_id") or doc.get("sku"),
            "existing_listing_id": existing_listing_id,
        }

    detail = _build_etsy_publish_comparison(doc)
    seo_optimization = await _generate_etsy_seo_optimization(detail)
    detail["etsy_optimized"]["title"] = seo_optimization.get("seo_title") or detail["etsy_optimized"].get("title")
    detail["etsy_optimized"]["when_made"] = seo_optimization.get("when_made") or detail["etsy_optimized"].get("when_made")

    payload_override = body.get("payload") if isinstance(body, dict) else None
    if payload_override is not None and not isinstance(payload_override, dict):
        raise HTTPException(status_code=400, detail="payload must be an object")

    request_payload = _build_etsy_create_listing_payload(detail, payload_override=payload_override)

    token = await get_valid_etsy_token()
    api_key = _resolve_etsy_api_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing Etsy API key")

    shop_id = await _resolve_etsy_shop_id()
    if not shop_id:
        raise HTTPException(status_code=400, detail="Unable to resolve Etsy shop_id")

    request_payload["readiness_state_id"] = await _resolve_etsy_readiness_state_id(
        shop_id=str(shop_id),
        payload=request_payload,
    )

    missing_fields = _get_missing_etsy_required_fields(request_payload)
    if request_payload.get("type") == "physical" and _to_int(request_payload.get("readiness_state_id")) is None:
        missing_fields.append("readiness_state_id")
    if missing_fields:
        failure_body = {
            "ok": False,
            "error": "missing_required_fields",
            "sku": detail.get("sku"),
            "missing_fields": missing_fields,
            "request_payload": request_payload,
        }
        await _record_etsy_draft_attempt(
            sku=str(detail.get("sku") or sku),
            shop_id=str(shop_id),
            request_payload=request_payload,
            request_form_data=None,
            response_status_code=400,
            response_payload=failure_body,
            ok=False,
            linked_listing_id=None,
            error_code="missing_required_fields",
        )
        return failure_body

    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": str(api_key),
        "Accept": "application/json",
    }

    form_data = _to_etsy_form_data(request_payload)
    url = f"{ETSY_BASE_URL}/shops/{shop_id}/listings"

    async with httpx.AsyncClient(timeout=45.0) as client:
        response = await client.post(url, headers=headers, data=form_data)

    raw_text = response.text
    try:
        etsy_response: Any = response.json() if raw_text else {}
    except Exception:
        etsy_response = raw_text

    listing_id = None
    if isinstance(etsy_response, dict):
        listing_id = _to_int(etsy_response.get("listing_id"))

    await _record_etsy_draft_attempt(
        sku=str(detail.get("sku") or sku),
        shop_id=str(shop_id),
        request_payload=request_payload,
        request_form_data=form_data,
        response_status_code=response.status_code,
        response_payload=etsy_response,
        ok=response.status_code < 300,
        linked_listing_id=listing_id,
        error_code=None if response.status_code < 300 else "etsy_create_draft_failed",
    )

    if response.status_code < 300 and listing_id:
        inventory_sync: dict[str, Any] = {
            "attempted": True,
            "ok": False,
            "error": None,
        }
        image_upload: dict[str, Any] = {
            "attempted": False,
            "ok": True,
            "uploaded": 0,
            "failed": 0,
            "results": [],
        }

        try:
            inventory_headers = {
                "Authorization": f"Bearer {token}",
                "x-api-key": str(api_key),
                "Accept": "application/json",
                "Content-Type": "application/json",
            }
            await _update_etsy_listing_sku_for_review(
                listing_id=listing_id,
                new_sku=str(detail.get("sku") or sku),
                target_quantity=max(1, _to_int(request_payload.get("quantity")) or 1),
                headers=inventory_headers,
            )
            inventory_sync["ok"] = True
        except Exception as exc:
            inventory_sync["error"] = str(exc)

        image_urls = _clean_list(
            (detail.get("etsy_optimized") or {}).get("images")
            or doc.get("images")
            or [],
            limit=10,
        )
        image_upload = await _upload_etsy_listing_images(
            shop_id=shop_id,
            listing_id=listing_id,
            image_urls=image_urls,
            headers=headers,
        )

        await db.product_normalized.update_one(
            {"_id": detail.get("sku")},
            {
                "$set": {
                    "channels.etsy.listing_id": listing_id,
                    "channels.etsy.shop_id": _to_int(shop_id) or shop_id,
                    "channels.etsy.listing_state": "draft",
                    "channels.etsy.title": request_payload.get("title"),
                    "channels.etsy.price": request_payload.get("price"),
                    "channels.etsy.quantity": request_payload.get("quantity"),
                    "channels.etsy.type": request_payload.get("type"),
                    "channels.etsy.who_made": request_payload.get("who_made"),
                    "channels.etsy.when_made": request_payload.get("when_made"),
                    "channels.etsy.taxonomy_id": request_payload.get("taxonomy_id"),
                    "channels.etsy.shipping_profile_id": request_payload.get("shipping_profile_id"),
                    "channels.etsy.return_policy_id": request_payload.get("return_policy_id"),
                    "channels.etsy.readiness_state_id": request_payload.get("readiness_state_id"),
                    "channels.etsy.last_create_draft_request": request_payload,
                    "channels.etsy.last_create_draft_response": etsy_response,
                    "channels.etsy.last_create_draft_http_status": response.status_code,
                    "channels.etsy.last_create_draft_ok": True,
                    "channels.etsy.last_create_draft_at": _now_utc(),
                    "channels.etsy.last_create_draft_inventory_sync": inventory_sync,
                    "channels.etsy.last_create_draft_image_upload": image_upload,
                }
            },
        )

        return {
            "ok": response.status_code < 300,
            "sku": detail.get("sku"),
            "shop_id": shop_id,
            "etsy_status_code": response.status_code,
            "etsy_response": etsy_response,
            "request_payload": request_payload,
            "request_form_data": form_data,
            "linked_listing_id": listing_id,
            "inventory_sync": inventory_sync,
            "image_upload": image_upload,
        }

    return {
        "ok": response.status_code < 300,
        "sku": detail.get("sku"),
        "shop_id": shop_id,
        "etsy_status_code": response.status_code,
        "etsy_response": etsy_response,
        "request_payload": request_payload,
        "request_form_data": form_data,
        "linked_listing_id": listing_id,
    }


@router.get("/channel-compare/list")
async def channel_compare_list(
    q: str | None = Query(default=None, description="Search SKU/title"),
    drift_only: bool = Query(default=False, description="Only rows with detected drift"),
    limit: int = Query(default=100, ge=1, le=500),
):
    query: dict = {"channels.etsy.listing_id": {"$exists": True}}
    if q:
        rx = _rx(q)
        query["$or"] = [
            {"_id": rx},
            {"sku": rx},
            {"title": rx},
            {"channels.etsy.title": rx},
        ]

    cursor = db.product_normalized.find(
        query,
        {
            "_id": 1,
            "sku": 1,
            "title": 1,
            "price": 1,
            "quantity": 1,
            "images": 1,
            "shopify_id": 1,
            "channels.etsy": 1,
            "channels.shopify": 1,
            "last_normalized_at": 1,
        },
    ).sort("last_normalized_at", -1)

    docs = await cursor.to_list(length=limit * 3)
    items: list[dict] = []
    for doc in docs:
        channels = doc.get("channels") or {}
        etsy = channels.get("etsy") or {}
        shopify = channels.get("shopify") or {}
        shopify_id = shopify.get("shopify_id") or doc.get("shopify_id")

        drift_fields: list[str] = []
        if _cmp_values(doc.get("title"), etsy.get("title")) == "diff":
            drift_fields.append("title")
        if _cmp_values(doc.get("price"), etsy.get("price"), numeric=True) == "diff":
            drift_fields.append("price")
        if _cmp_values(doc.get("quantity"), etsy.get("quantity"), numeric=True) == "diff":
            drift_fields.append("quantity")
        if not shopify_id:
            drift_fields.append("shopify_missing")

        if drift_only and not drift_fields:
            continue

        items.append(
            {
                "sku": doc.get("_id") or doc.get("sku"),
                "title": doc.get("title"),
                "image": (doc.get("images") or [None])[0],
                "etsy_listing_id": etsy.get("listing_id"),
                "shopify_id": shopify_id,
                "match_bucket": ((etsy.get("match") or {}).get("bucket")),
                "drift_fields": drift_fields,
                "drift_count": len(drift_fields),
            }
        )
        if len(items) >= limit:
            break

    return {
        "count": len(items),
        "items": items,
    }


@router.get("/channel-compare/kpis")
async def channel_compare_kpis():
    total_products = await db.product_normalized.count_documents({})

    etsy_linked = await db.product_normalized.count_documents(
        {"channels.etsy.listing_id": {"$exists": True, "$ne": None}}
    )
    shopify_linked = await db.product_normalized.count_documents(
        {"$or": [
            {"channels.shopify.shopify_id": {"$exists": True, "$ne": None}},
            {"shopify_id": {"$exists": True, "$ne": None}},
        ]}
    )

    inventory_pipeline = [
        {
            "$group": {
                "_id": None,
                "ebay_total_inventory": {
                    "$sum": {"$ifNull": ["$quantity", 0]}
                },
                "etsy_total_inventory": {
                    "$sum": {"$ifNull": ["$channels.etsy.quantity", 0]}
                },
                "shopify_total_inventory": {
                    "$sum": {"$ifNull": ["$channels.shopify.quantity", 0]}
                },
            }
        }
    ]
    agg = await db.product_normalized.aggregate(inventory_pipeline).to_list(length=1)
    totals = agg[0] if agg else {}

    return {
        "kpis": {
            "ebay_total_inventory": int(totals.get("ebay_total_inventory", 0) or 0),
            "etsy_total_inventory": int(totals.get("etsy_total_inventory", 0) or 0),
            "shopify_total_inventory": int(totals.get("shopify_total_inventory", 0) or 0),
            "total_products": int(total_products),
            "etsy_linked_products": int(etsy_linked),
            "shopify_linked_products": int(shopify_linked),
            "etsy_missing_products": max(0, int(total_products) - int(etsy_linked)),
            "shopify_missing_products": max(0, int(total_products) - int(shopify_linked)),
        }
    }


@router.get("/channel-compare/{sku}")
async def channel_compare_detail(sku: str):
    doc = await db.product_normalized.find_one(
        {"$or": [{"_id": sku}, {"sku": sku}]},
        {
            "_id": 1,
            "sku": 1,
            "title": 1,
            "price": 1,
            "quantity": 1,
            "images": 1,
            "shopify_id": 1,
            "channels.etsy": 1,
            "channels.shopify": 1,
            "last_normalized_at": 1,
        },
    )
    if not doc:
        raise HTTPException(status_code=404, detail="SKU not found")

    channels = doc.get("channels") or {}
    etsy = channels.get("etsy") or {}
    shopify_channel = channels.get("shopify") or {}
    shopify_id = shopify_channel.get("shopify_id") or doc.get("shopify_id")
    shopify_live = await _fetch_shopify_snapshot(shopify_id)

    normalized_price = _to_float(doc.get("price"))
    normalized_qty = _to_int(doc.get("quantity"))
    etsy_price = _to_float(etsy.get("price"))
    etsy_qty = _to_int(etsy.get("quantity"))
    shopify_price = _to_float((shopify_live or {}).get("price"))
    shopify_qty = _to_int((shopify_live or {}).get("quantity"))

    comparisons = {
        "title": {
            "etsy_vs_shopify": _cmp_values(etsy.get("title"), (shopify_live or {}).get("title")),
            "normalized_vs_etsy": _cmp_values(doc.get("title"), etsy.get("title")),
            "normalized_vs_shopify": _cmp_values(doc.get("title"), (shopify_live or {}).get("title")),
        },
        "price": {
            "etsy_vs_shopify": _cmp_values(etsy_price, shopify_price, numeric=True),
            "normalized_vs_etsy": _cmp_values(normalized_price, etsy_price, numeric=True),
            "normalized_vs_shopify": _cmp_values(normalized_price, shopify_price, numeric=True),
        },
        "quantity": {
            "etsy_vs_shopify": _cmp_values(etsy_qty, shopify_qty, numeric=True),
            "normalized_vs_etsy": _cmp_values(normalized_qty, etsy_qty, numeric=True),
            "normalized_vs_shopify": _cmp_values(normalized_qty, shopify_qty, numeric=True),
        },
        "status": {
            "etsy_vs_shopify": _cmp_values(etsy.get("listing_state"), (shopify_live or {}).get("status")),
        },
    }

    return {
        "sku": doc.get("_id") or doc.get("sku"),
        "normalized": {
            "title": doc.get("title"),
            "price": normalized_price,
            "quantity": normalized_qty,
            "image": (doc.get("images") or [None])[0],
            "last_normalized_at": doc.get("last_normalized_at"),
        },
        "etsy": {
            "listing_id": etsy.get("listing_id"),
            "title": etsy.get("title"),
            "price": etsy_price,
            "quantity": etsy_qty,
            "status": etsy.get("listing_state"),
            "url": etsy.get("url") or (
                f"https://www.etsy.com/listing/{etsy.get('listing_id')}" if etsy.get("listing_id") else None
            ),
            "image": etsy.get("image"),
            "match": etsy.get("match") or {},
        },
        "shopify": {
            "shopify_id": shopify_id,
            "links": _build_shopify_links(shopify_id),
            "live": shopify_live,
        },
        "comparisons": comparisons,
    }


@router.post("/etsy-publish/bulk/validate")
async def etsy_publish_bulk_validate(
    body: dict | None = Body(default=None),
    min_taxonomy_confidence: float = Query(default=ETSY_BULK_MIN_TAXONOMY_CONFIDENCE, ge=0.0, le=1.0),
):
    """
    Validate items for bulk publishing (dry-run).
    
    Request body:
    {
        "skus": ["sku1", "sku2", ...],  // Optional; if omitted, validates all unlinked items
        "all": true  // Optional; if true, validate all unlinked items
    }
    
    Returns validation results with per-item details including skip reasons.
    Uses async/await with rate limiting to respect OpenAI and Etsy API limits.
    """
    request_skus = (body or {}).get("skus") if isinstance(body, dict) else None
    request_all = (body or {}).get("all") if isinstance(body, dict) else None
    
    # Determine which SKUs to validate
    if isinstance(request_skus, list) and request_skus:
        # Validate specified SKUs
        query = {
            "quantity": {"$gte": 1},
            "$or": [{"_id": {"$in": request_skus}}, {"sku": {"$in": request_skus}}],
        }
    elif request_all:
        # Validate all unlinked items with qty >= 1
        query = {
            "quantity": {"$gte": 1},
            "$or": [
                {"channels.etsy.listing_id": {"$exists": False}},
                {"channels.etsy.listing_id": None},
                {"channels.etsy.listing_id": ""},
            ],
        }
    else:
        raise HTTPException(status_code=400, detail="Provide either 'skus' list or 'all': true in request body")
    
    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }
    
    docs = await db.product_normalized.find(query, projection).sort("last_normalized_at", -1).to_list(length=ETSY_BULK_MAX_ITEMS_PER_RUN)
    
    # Limit endpoint-level concurrency to keep the API responsive while validating.
    validation_results: list[dict[str, Any]] = []
    batch_size = max(1, ETSY_BULK_VALIDATE_CONCURRENCY)

    for start in range(0, len(docs), batch_size):
        batch = docs[start:start + batch_size]
        batch_results = await asyncio.gather(
            *[
                _validate_bulk_item(doc, min_taxonomy_confidence=min_taxonomy_confidence)
                for doc in batch
            ],
            return_exceptions=True,
        )

        for idx, result in enumerate(batch_results):
            if isinstance(result, Exception):
                failed_doc = batch[idx]
                failed_sku = failed_doc.get("_id") or failed_doc.get("sku")
                validation_results.append(
                    {
                        "sku": str(failed_sku),
                        "status": "skipped",
                        "reason": "validation_exception",
                        "detail": str(result),
                        "optimizations": None,
                    }
                )
            elif isinstance(result, dict):
                validation_results.append(result)

        # Yield to the event loop between batches so other endpoints can run.
        await asyncio.sleep(0)
        if ETSY_BULK_VALIDATE_BATCH_DELAY_SECONDS > 0:
            await asyncio.sleep(ETSY_BULK_VALIDATE_BATCH_DELAY_SECONDS)
    
    report = await _generate_bulk_report(validation_results)
    session_id = f"validate_{_now_utc().timestamp()}"
    validated_at = _now_utc()
    _ETSY_BULK_REPORT["session_id"] = session_id
    _ETSY_BULK_REPORT["validated_at"] = validated_at
    _ETSY_BULK_REPORT["validation_result"] = report
    _ETSY_BULK_REPORT["creation_result"] = None

    await _persist_bulk_validation_report(
        session_id=session_id,
        validated_at=validated_at,
        validation_result=report,
    )

    report = dict(report)
    report["session_id"] = session_id
    report["validated_at"] = validated_at
    
    return report


@router.post("/etsy-publish/bulk/create")
async def etsy_publish_bulk_create(
    body: dict | None = Body(default=None),
):
    """
    Execute bulk publishing of validated items using pre-computed optimizations.
    
    Request body:
    {
        "skus": ["sku1", "sku2", ...],  // Optional; uses validated ready items if omitted
        "confirmed": true  // Required; must be true to actually create (safety check)
    }
    
    Uses optimizations cached from the last validation run. Much faster than create-draft
    since all AI optimization work was done during validation.
    
    Returns creation results with per-item success/failure details.
    Uses async/await with rate limiting to respect Etsy API limits (10 req/sec).
    """
    request_skus = (body or {}).get("skus") if isinstance(body, dict) else None
    request_session_id = (body or {}).get("session_id") if isinstance(body, dict) else None
    confirmed = (body or {}).get("confirmed") if isinstance(body, dict) else False
    
    if not confirmed:
        raise HTTPException(status_code=400, detail="Must set 'confirmed': true to proceed")
    
    # Prefer persisted report so create works across restarts/workers.
    persisted_report = await _fetch_bulk_report(session_id=str(request_session_id) if request_session_id else None)
    active_session_id = (persisted_report or {}).get("session_id") or _ETSY_BULK_REPORT.get("session_id")
    validation_result = (persisted_report or {}).get("validation_result") or _ETSY_BULK_REPORT.get("validation_result") or {}

    # If no specific SKUs provided, use validated ready items from last validation
    if not isinstance(request_skus, list) or not request_skus:
        ready_items = validation_result.get("validation_items") or []
        request_skus = [item["sku"] for item in ready_items if item.get("status") == "ready"]
        if not request_skus:
            raise HTTPException(
                status_code=400,
                detail="No ready items in last validation. Run validation first with all: true",
            )
    
    # Build optimization cache from validation results (sku -> optimizations)
    validation_items = validation_result.get("validation_items") or []
    optimizations_cache: dict[str, dict] = {}
    for item in validation_items:
        if item.get("status") == "ready" and item.get("optimizations"):
            optimizations_cache[str(item["sku"])] = item["optimizations"]
    
    # Fetch documents
    query = {
        "$or": [{"_id": {"$in": request_skus}}, {"sku": {"$in": request_skus}}],
    }
    projection = {
        "_id": 1,
        "sku": 1,
        "title": 1,
        "description": 1,
        "price": 1,
        "quantity": 1,
        "tags": 1,
        "images": 1,
        "attributes": 1,
        "category": 1,
        "package": 1,
        "shipping": 1,
        "channels.etsy": 1,
        "last_normalized_at": 1,
    }
    docs_dict = {}
    async for doc in db.product_normalized.find(query, projection):
        sku = doc.get("_id") or doc.get("sku")
        docs_dict[str(sku)] = doc
    
    # Create single HTTP client for reuse (more efficient than creating per request)
    token = await get_valid_etsy_token()
    api_key = _resolve_etsy_api_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing Etsy API key")
    
    shop_id = await _resolve_etsy_shop_id()
    if not shop_id:
        raise HTTPException(status_code=400, detail="Missing shop_id")
    
    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": str(api_key),
        "Accept": "application/json",
    }
    
    # Async function to create a single listing with Etsy rate limiting
    async def create_single_listing(sku: str, index: int) -> dict[str, Any]:
        doc = docs_dict.get(str(sku))
        if not doc:
            return {
                "sku": str(sku),
                "status": "failed",
                "listing_id": None,
                "error": "document_not_found",
                "etsy_status_code": None,
            }
        
        # Get cached optimizations for this SKU
        cached_opt = optimizations_cache.get(str(sku))
        if not cached_opt:
            return {
                "sku": str(sku),
                "status": "failed",
                "listing_id": None,
                "error": "no_cached_optimizations",
                "etsy_status_code": None,
            }
        
        try:
            # Build detail and apply cached optimizations
            detail = _build_etsy_publish_comparison(doc)
            
            # Apply cached optimizations directly
            detail["etsy_optimized"]["title"] = cached_opt.get("seo_title") or detail["etsy_optimized"].get("title")
            detail["etsy_optimized"]["when_made"] = cached_opt.get("when_made") or detail["etsy_optimized"].get("when_made")
            detail["etsy_optimized"]["tags"] = cached_opt.get("tags") or detail["etsy_optimized"].get("tags")
            detail["etsy_optimized"]["taxonomy_id"] = cached_opt.get("taxonomy_id") or detail["etsy_optimized"].get("taxonomy_id")
            
            request_payload = _build_etsy_create_listing_payload(detail)
            
            request_payload["readiness_state_id"] = await _resolve_etsy_readiness_state_id(
                shop_id=str(shop_id),
                payload=request_payload,
            )
            
            missing_fields = _get_missing_etsy_required_fields(request_payload)
            if missing_fields:
                return {
                    "sku": str(detail.get("sku") or sku),
                    "status": "failed",
                    "listing_id": None,
                    "error": "missing_required_fields",
                    "etsy_status_code": None,
                }
            
            form_data = _to_etsy_form_data(request_payload)
            url = f"{ETSY_BASE_URL}/shops/{shop_id}/listings"
            
            # Rate limit Etsy API requests using semaphore
            semaphore = _get_etsy_semaphore()
            async with semaphore:
                async with httpx.AsyncClient(timeout=45.0) as client:
                    response = await client.post(url, headers=headers, data=form_data)
                # Add delay after request to respect rate limit
                await asyncio.sleep(ETSY_REQUEST_DELAY_SECONDS)
            
            raw_text = response.text
            try:
                etsy_response: Any = response.json() if raw_text else {}
            except Exception:
                etsy_response = raw_text
            
            listing_id = None
            if isinstance(etsy_response, dict):
                listing_id = _to_int(etsy_response.get("listing_id"))
            
            success = response.status_code < 300
            
            await _record_etsy_draft_attempt(
                sku=str(detail.get("sku") or sku),
                shop_id=str(shop_id),
                request_payload=request_payload,
                request_form_data=form_data,
                response_status_code=response.status_code,
                response_payload=etsy_response,
                ok=success,
                linked_listing_id=listing_id,
                error_code=None if success else "etsy_create_draft_failed",
            )
            
            if success and listing_id:
                # Post-create: upload images and sync inventory
                image_urls = doc.get("images") or []
                if image_urls:
                    await _upload_etsy_listing_images(
                        shop_id=shop_id,
                        listing_id=listing_id,
                        image_urls=image_urls,
                        headers=headers,
                    )
                
                await _update_etsy_listing_sku_for_review(
                    listing_id=listing_id,
                    new_sku=str(detail.get("sku") or sku),
                    target_quantity=_to_int(doc.get("quantity")) or 0,
                    headers=headers,
                )
            
            return {
                "sku": str(detail.get("sku") or sku),
                "status": "created" if success else "failed",
                "listing_id": listing_id,
                "error": None if success else etsy_response.get("error", "unknown_error"),
                "etsy_status_code": response.status_code,
            }
        
        except Exception as exc:
            return {
                "sku": str(sku),
                "status": "failed",
                "listing_id": None,
                "error": str(exc),
                "etsy_status_code": None,
            }
    
    # Process all listings concurrently with Etsy rate limiting
    tasks = [create_single_listing(sku, idx) for idx, sku in enumerate(request_skus)]
    creation_results = await asyncio.gather(*tasks, return_exceptions=False)
    
    # Filter out any failed tasks
    creation_results = [r for r in creation_results if isinstance(r, dict)]
    
    report = await _generate_bulk_report([], creation_results)
    _ETSY_BULK_REPORT["creation_result"] = report
    created_at = _now_utc()
    _ETSY_BULK_REPORT["created_at"] = created_at

    if active_session_id:
        await _persist_bulk_creation_report(
            session_id=str(active_session_id),
            created_at=created_at,
            creation_result=report,
        )

    report = dict(report)
    report["session_id"] = active_session_id
    report["created_at"] = created_at
    
    return report


@router.get("/etsy-publish/bulk/last-report")
async def etsy_publish_bulk_last_report(
    session_id: str | None = Query(default=None, description="Fetch a specific session_id (optional)"),
    include_validation: bool = Query(default=True, description="Include validation items in report"),
    include_creation: bool = Query(default=True, description="Include creation items in report"),
):
    """
    Fetch the last bulk publishing report.
    
    Returns combined validation and creation results if available.
    """
    persisted = await _fetch_bulk_report(session_id=session_id)
    report = copy.deepcopy(persisted) if persisted else copy.deepcopy(_ETSY_BULK_REPORT)
    
    validation_result = report.get("validation_result") or {}
    creation_result = report.get("creation_result") or {}
    
    # Filter items based on query parameters
    if not include_validation:
        validation_result.pop("validation_items", None)
    if not include_creation:
        creation_result.pop("creation_items", None)
    
    return {
        "session_id": report.get("session_id"),
        "validated_at": report.get("validated_at"),
        "created_at": report.get("created_at"),
        "validation": validation_result,
        "creation": creation_result,
        "summary": {
            "validation_total": validation_result.get("validation", {}).get("total_checked"),
            "validation_ready": validation_result.get("validation", {}).get("ready_to_create"),
            "validation_skipped": validation_result.get("validation", {}).get("skipped"),
            "creation_total": creation_result.get("creation", {}).get("total_attempted"),
            "creation_created": creation_result.get("creation", {}).get("created"),
            "creation_failed": creation_result.get("creation", {}).get("failed"),
        }
    }
